import time
import os
import io
import pandas as pd
import numpy as np
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn import metrics
import time
import psutil
import autogluon
import pandas as pd
import numpy as np
import json
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from scipy import interp
from scipy.stats import norm
from scipy.special import ndtri
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import magic 
import pickle
import random 
from random import randint
from random import uniform
from scipy import stats
import ast
from pathlib import Path
import joblib as joblib
from joblib import dump, load
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from sklearn.model_selection import cross_validate
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import RocCurveDisplay
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.metrics import auc
from sklearn.utils import resample
from sklearn.model_selection import KFold
from sklearn.model_selection import StratifiedKFold, RepeatedStratifiedKFold
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import VarianceThreshold
from sklearn.model_selection import RandomizedSearchCV, GridSearchCV
#from tune_sklearn import TuneSearchCV
from skopt import BayesSearchCV
from sklearn.calibration import CalibratedClassifierCV
from sklearn.calibration import calibration_curve, CalibrationDisplay
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score
from sklearn.metrics import classification_report
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import label_binarize
from sklearn.metrics import confusion_matrix
from scipy import interp
from scipy.stats import norm
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
import random 
from random import randint
from random import uniform
from scipy import stats
import mrmr
from mrmr import mrmr_classif
import xgboost
import catboost
import shap
from scipy import stats
import os
import joblib as joblib
from joblib import dump, load
import json
import tkinter as tk
from tkinter import *
from autogluon.tabular import TabularDataset, TabularPredictor
#np.random.seed(1000)
from difflib import SequenceMatcher
#np.random.seed(1000)
rstate = 12

# import module
import datetime
import pprint
import pymongo
from pymongo import MongoClient
import streamlit as st

# connect to database
client = MongoClient('10.14.1.12', 27017)

# create the database if it does not already exists
db = client.machine_learning_database

# create tables for models in the databse
models = db.models

from Multi_Outcome_Classification_tools import multi_outcome_hyperparameter_binary
from Common_Tools import wrap_text_excel, expand_cell_excel, grid_excel, generate_all_idx_files, upload_data, load_data, save_data, data_prep, data_prep_train_set, parse_exp_multi_outcomes, setup_multioutcome_binary, refine_binary_outcomes, generate_joblib_model
from roctools import full_roc_curve, plot_roc_curve

global algo_shortnames # short names for ML Algorithims
algo_shortnames = {
    "Random Forest": "rf",
    "XGBoost": "xgb",
    "Cat Boost": "cat",
    "SGD Elastic": "sgd_elastic",
    "SGD L2": "sgd_l2",
    "Logistic Reg. L2": "lr_l2",
    "Logistic Reg.": "lr", 
    "Descision Tree": "dt", 
    "SVM": "svm",
    "KNearest N.": "knn", 
}

data_sets = {}

def run_all_experiments(all_experiments, project_folder, data_sets, unique_value_threshold=10):

    project_name = list(all_experiments.keys())[0]
    #experiment_names = list(all_experiments[project_name].keys())
    #st.write("List of experiments: ", experiment_names)

    train_set = data_sets["Training Set"]
    index_set = data_sets["Index Set"]
    test_sets = data_sets["Testing Set"]

    # get input and ouput columns
    input_cols, label_cols, categorical_cols, numeric_cols = parse_exp_multi_outcomes(train_set["Data"], index_set["Data"], unique_value_threshold=unique_value_threshold)
    
    #st.write("Input Columns: ", input_cols)
    st.write("Label Columns: ", label_cols)

    problem_type = all_experiments[project_name]['exp_type']
    #st.write(problem_type)

    # refine the output data
    if problem_type == "binary":
        df_train = refine_binary_outcomes(train_set["Data"], label_cols)
    elif problem_type == "multiclass":
        print("Multiclass")
        #df_train = refine_multiclass_outcomes(train_set["Data"], label_cols)
    
    algorithms = [] # list of algorithims

    for experiment in all_experiments[project_name]:
        experiment_name = experiment
        if experiment_name in ["train_set", "test_sets", "exp_type"]:
            continue

        st.write(f"Experiment Name: {experiment_name} had started training")

        set_up = all_experiments[project_name][experiment_name]

        if problem_type == "binary":
            print("Binary")
            options, algorithm, param_vals = setup_multioutcome_binary(set_up, experiment_name, project_folder)
        elif problem_type == "multiclass":
            print("Multiclass")
            options, algorithm, param_vals = setup_multioutcome_multiclass(set_up, experiment_name, project_folder)

        algorithms.append(algorithm)

         # refine all the test sets and save them 
        for _, (testing_set_name, testing_set) in enumerate(data_sets['Testing Set'].items()):

            # prep testing sets
            _, df_test, _, _ = data_prep_train_set(df_train, testing_set, input_cols, label_cols, numeric_cols, categorical_cols, options)

            # save test set refined
            file_name = "refined_" + testing_set_name
            df_test.to_csv(os.path.join(project_folder, experiment_name, file_name))

        # refine the training set before model training
        df_train_refined, input_cols, label_cols = data_prep(df_train, input_cols, label_cols, numeric_cols, categorical_cols, options)
        #st.write(df_train_refined)

        #st.write('START TRAINING (FINALLY!!!!)') 
        multi_outcome_hyperparameter_binary(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, options, algorithm, param_vals, experiment_name, project_folder)
        #if training_type == 'Traditional':
            #st.write('Traditional')
            #multi_outcome_traditional_binary(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, options, algorithm, param_vals, experiment_name, project_folder)
        #elif training_type == 'Hyperparameter Tuning':
            #st.write('Hyperparameter Tuning')
            #multi_outcome_hyperparameter_binary(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, options, algorithm, param_vals, experiment_name, project_folder)
        #else:
            #st.write("No model has been trained.") 

          
    models_dic, input_cols_dic = generate_joblib_model(project_folder)
    model_absolute_path = os.path.join(project_folder)

    pathway_name = os.path.join(model_absolute_path, project_name + "_models.joblib")
    joblib.dump(models_dic, pathway_name)
    
    model = {
        "exp_name": project_name,
        "type": "Native",
        "model_path": pathway_name,
        "algorithms": algorithms,
        "input variables": input_cols_dic,
        "configuration": all_experiments
    }

    models.insert_one(model) # insert one dictonary 




def project(all_experiments, data_sets, unique_value_threshold=10):
    
    project_name = list(all_experiments.keys())[0]
    st.write(project_name)
    
    project_folder = os.path.join("Models", project_name)
    os.makedirs(project_folder, exist_ok=True)   # Create folder for experiment results

    # Print setup
    #st.write(all_experiments[project_name])
    
    # Pause program to manually set up the index files
    #print("Set up the index files, 1 = input, 2 = output")
    #wait = input("Press Enter to continue.")
    #print("Continue")
    
    '''
    # Training the models...
    '''
    # run all experiments
    #try:
    run_all_experiments(all_experiments, project_folder, data_sets, unique_value_threshold=unique_value_threshold)
    #except:
        #st.write("An Error Has Occured in Training.")
    
    if st.button("Test the Models"):
        st.switch_page("pages/Testing_Native_Models.py")  # Redirect to visualize_results.py

    #if type == 'Traditional': # if all experiments are binary classification
    #    binary_run_all_experiments(all_experiments, project_name, unique_value_threshold=unique_value_threshold)
    #elif type == 'Hyperparamter Tuning': # if all experiments are multiclass classification
    #    hpt_run_all_experiments(all_experiments, project_name, unique_value_threshold=unique_value_threshold)
    #else:
    #    st.write("No model has been trained.")



def gen_idx(df_train, train_set_name, nexp:int=1, sheets:list=None):

    if train_set_name.endswith('.xlsx'):
        filename = train_set_name[:-len('.xlsx')]
    elif train_set_name.endswith('.csv'):
        filename = train_set_name[:-len('.csv')]
    else:
        filename = train_set_name

    indexFileName = filename + '_index.xlsx'
    #st.write(indexFileName)
    #writer = pd.ExcelWriter(indexFileName, engine='xlsxwriter')
    
    
    all_columns = list(df_train.columns)

    df_index = pd.DataFrame(columns=all_columns)
    df_index.loc['Key'] = 0

    #st.write(df_index)

    # convert df_index to an Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_index.to_excel(writer, index=True)

    st.download_button(
        label="Download the Index File Template to set inputs and outputs",
        data=output.getvalue(),
        file_name=indexFileName,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return indexFileName

def generate_congfig_file(exp_name, algorithims, exp_type, options):
    print("Number of exp. :", len(algorithims))

    configuration_dic = {}
    configuration_dic[exp_name] = {}
    #configuration_dic[exp_name]['train_set']  = train_set
    #configuration_dic[exp_name]['test_sets']  = test_sets
    configuration_dic[exp_name]['exp_type']  = exp_type

    for algorithim in algorithims:
        experiment = {}
        experiment['algorithm'] = algo_shortnames[algorithim]
        #experiment['training_type'] = training_type
        experiment['options'] = options
        experiment['param_vals'] = "None"

        configuration_dic[exp_name][f"exp_{algorithim}"] = experiment
    
    #st.write("Configuration_dic : ", configuration_dic)
    
    return configuration_dic
    
# back button to return to main page
if st.button('Back'):
    st.switch_page("pages/Training_Models_Native_Options.py")  # Redirect to the main back

# Title
st.title("Train and Develop ML Model (Native)")

st.write("This page allows you to train and develop your machine learning model for standard (native/sklearn) types.")

# get all unique exp. names from results collection
exp_names = db.models.distinct("exp_name", {"type": "Native"})

# choose how you configure your model(s)
configure_options = st.radio("Choose an option:", ["Upload a file", "User Customization"])

if configure_options == "Upload a file":
    # reset configuration_dic
    configuration_dic = None

    # sends user to a seperate page where they can download an empty configuration file where they can customize themselves
    if st.button('Download a Configuration File Template'):
        st.switch_page("pages/Configuration_Page.py")  # Redirect to Configuration_Page.py

    # File uploader for the training
    uploaded_config_file= st.file_uploader("Upload a Configuration File")

    if uploaded_config_file:
        #st.write(uploaded_config_file)

        # open JSON object as 
        # a dictionary
        configuration_dic = json.load(uploaded_config_file)

        project_name = list(configuration_dic.keys())[0] # get the name of model configuration and check if the model already exists in the database

        if project_name in exp_names:
            st.write("Model with that name already exists")
            already_exists = True
        else:
            already_exists = False

        #st.write(configuration_dic)
else:
    # enter experiment name
    exp_name = st.text_input("Enter Name of the ML Experiment", "exp_name")

    if exp_name in exp_names:
        st.write("Model with that name already exists")
        already_exists = True
    else:
        already_exists = False
# Number input
#num_models = st.number_input("Enter the number of models:", min_value=1, max_value=100, value=5)

# File uploader for the training
uploaded_train_set= st.file_uploader("Upload a Training Data Set (Only one datset)")

if uploaded_train_set:
    train_set = uploaded_train_set.name
    st.write(train_set)

    df_train = load_data(train_set, uploaded_train_set)

    #df_train = pd.read_csv(uploaded_train_set)
    st.write(df_train.head())  # Display the first few rows

    data_sets["Training Set"] = {}
    data_sets["Training Set"]["Name"] = uploaded_train_set.name
    data_sets["Training Set"]["Data"] = df_train

    gen_idx(df_train, uploaded_train_set.name)

# File uploader for the index table
uploaded_index_set = st.file_uploader("Upload a Completed Index Table")

if uploaded_index_set:

    index_set = pd.read_excel(uploaded_index_set)
    st.write(index_set.head())  # Display the first few rows

    data_sets["Index Set"] = {}
    data_sets["Index Set"]["Name"] = uploaded_index_set.name
    data_sets["Index Set"]["Data"] = index_set

# File uploader for the test
uploaded_test_set= st.file_uploader("Upload a Testing Data Set (Can Upload Multiple Datasets)", accept_multiple_files=True)

test_sets = []

if uploaded_test_set:

    data_sets["Testing Set"] = {}

    for file in uploaded_test_set:
        st.subheader(f"Dataset: {file.name}")
        test_sets.append(file.name)

        df_test = load_data(file.name, file)

        st.write(df_test.head())  # Display the first few rows

        data_sets["Testing Set"][file.name] = df_test

st.write(test_sets)

if configure_options == "User Customization":
    # reset configuration_dic
    configuration_dic = None

    # enter the list of ML algorithims
    algo_list = ['Random Forest', 'XGBoost', 'Cat Boost', "SGD Elastic", "SGD L2", "Logistic Reg.", "Logistic Reg. L2", "Descision Tree", "SVM", "KNearest N."]
    algorithims = st.multiselect("ML Algorithims to Use (Must have at least 1)", algo_list)

    # enter the type of ML training
    #training_type = st.selectbox("Select the training method", ['Select one', 'Traditional', 'Hyperparameter Tuning'])
else:
    algorithims = []

if configure_options == "User Customization" and len(algorithims) > 0 and already_exists == False:

    # one hot encoding
    oneHotEncode = st.radio("oneHotEncode:", ["True", "False"])

    # impute
    impute = st.radio("Impute:", ["True", "False"])

    # cutMissingRows
    cutMissingRows = st.radio("cutMissingRows:", ["True", "False"])

    # cut threshold
    if cutMissingRows == "True":
        cut_threshold = st.number_input("Enter the threshold for cutting missing values (min '%' of non-null values that a row must have to not be cut):", min_value=0.0, max_value=1.0, value=0.6)
    else:
        cut_threshold =  None

    # what to replace infinte values with
    inf = st.selectbox("Replace infinte values", ['replace with null', 'replace with zero'])

    # how to deal with outliers
    outliers = st.selectbox("Replace outlier values", ['remove rows', 'log'])
    outliers_N = st.number_input("Outlier threshold:", value=50000)

    # scaling
    scaling = st.radio("Scaling:", ["True", "False"])
    if scaling == "True":
        scalingMethod = st.selectbox("Select the scaling method", ['MinMaxScaler', 'RobustScaler', 'MaxAbsScaler', 'StandardScaler'])
    else:
        scalingMethod = None

    # QuantileTransformer
    QuantileTransformer = st.radio("QuantileTransformer:", ["True", "False"])

    # Normalize
    normalize = st.radio("Normalize:", ["True", "False"])

    # Rebalance
    rebalance = st.radio("Rebalance:", ["True", "False"])
    if rebalance == "True":
        rebalance_type = st.selectbox("Select the rebalancing method", ['RandomUnderSampler', 'RandomOverSampler', 'SMOTE', 'ADASYN'])
    else:
        rebalance_type = None

    # FeatureSelection
    featureSelection = st.radio("FeatureSelection:", ["True", "False"])
    if featureSelection == "True":
        featureSelection_method = st.selectbox("Select the Feature Selection method", ['MRMR', 'SelectKBest-f_classif', 'SelectKBest-chi2', 'SelectPercentile-f_classif', 'SelectPercentile-chi2', 'VarianceThreshold'])
        N_features = st.number_input("Enter the number of features to select:", min_value=1, max_value=100, value=5)
    else:
        featureSelection_method = None
        N_features = None
    

    # searching strategy
    search_strategy = st.selectbox("Select the searching strategy method", ['random', 'bayesian', 'grid', 'ray'])
    num_itr = st.number_input("Enter the number of iterations for searching:", min_value=1, value=50)

    # k folds for cross validation
    k_fold = st.number_input("Enter the number of K folds for cross validation:", min_value=2, value=10)
    n_repeats = st.number_input("Enter the number of repeats for cross validation:", min_value=0, value=1)

    # min number of postive values for an outcome to be trained
    min_postives = st.number_input("Enter the minimum number :", min_value=1, value=10)

    options = {
        'oneHotEncode' : oneHotEncode, 
        'Impute': impute, 
        'cutMissingRows' : cutMissingRows,
        "cut threshold": cut_threshold,
        "inf": inf,
        'outliers': outliers, 
        'outliers_N': outliers_N, 
        'Scaling': scaling, 
        'scalingMethod': scalingMethod, 
        'QuantileTransformer': QuantileTransformer, 
        'Normalize': normalize,
        'rebalance' : rebalance,
        'rebalance_type': rebalance_type,
        'FeatureSelection': featureSelection,
        "method": featureSelection_method,
        'N_features': N_features, 
        'strategy': search_strategy, 
        'itr': num_itr,
        "CV": k_fold,
        "n_repeats": n_repeats,
        "min_postives": min_postives,
    }

    configuration_dic = generate_congfig_file(exp_name, algorithims, "binary", options)

        

if configuration_dic and uploaded_train_set and uploaded_index_set and uploaded_test_set and already_exists == False and st.button("Start Training the Models"):
    #st.write(configuration_dic)

    # get name of the project as part of file name
    project_name = list(configuration_dic.keys())[0]

    #st.write(data_sets)
    project_folder = os.path.join("Models", project_name)
    os.makedirs(project_folder, exist_ok=True)  # Create folder for algorithm results

    save_data(data_sets['Training Set']['Name'], data_sets['Training Set']['Data'], os.path.join(project_folder, data_sets['Training Set']['Name']))

    save_data(data_sets['Index Set']['Name'], data_sets['Index Set']['Data'], os.path.join(project_folder, data_sets['Index Set']['Name']))
    #for _, (testing_set_name, testing_set) in enumerate(data_sets['Testing Set'].items()):
    
        
    project(configuration_dic, data_sets, unique_value_threshold=10)
