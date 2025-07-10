import pandas as pd
import numpy as np
import inspect, re
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.model_selection import cross_validate
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import RocCurveDisplay
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.metrics import auc
from sklearn.utils import resample
from sklearn.model_selection import KFold
from sklearn.model_selection import StratifiedKFold
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import VarianceThreshold
from sklearn.model_selection import RandomizedSearchCV, GridSearchCV
from skopt import BayesSearchCV
from sklearn.calibration import CalibratedClassifierCV
#from tune_sklearn import TuneSearchCV
from sklearn.calibration import calibration_curve, CalibrationDisplay
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import classification_report
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import label_binarize
from sklearn.metrics import confusion_matrix
from scipy import interp
from scipy.stats import norm
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import xlsxwriter
import magic 
import random 
from random import randint
from random import uniform
from scipy import stats
import mrmr
from mrmr import mrmr_classif
import xgboost
import catboost
import shap
from scipy import stats
import os
import hashlib
import joblib as joblib
from joblib import dump, load
import json
import tkinter as tk
from tkinter import *
import time
import psutil
from PIL import ImageTk, Image
import statistics
from IPython.display import display
import datetime
import pprint
import pymongo
from pymongo import MongoClient
import streamlit as st
#np.random.seed(1000)
rstate = 12

#import Multiclass_Classification_tools
#import Binary_Classification_tools

global options_default
global options_test_set
global algo_shortnames # short names for ML Algorithims
algo_shortnames = {
    "Random Forest": "rf",
    "XGBoost": "xgb",
    "Cat Boost": "cat",
    "SGD Elastic": "sgd_elastic",
    "SGD L2": "sgd_l2",
    "Logistic Reg. L2": "lr_l2",
    "Logistic Reg.": "lr", 
    "Descision Tree": "dt", 
    "SVM": "svm",
    "KNearest N.": "knn", 
}

options_test_set = { # default options dict
        'oneHotEncode' : "True", 
        'Impute': "True", 
        'cutMissingRows' : "True",
        "cut threshold": 0.60,
        "inf": 'replace with null',
        'outliers': "log",
        'outliers_N': 50000, 
        'Scaling': "True", 
        'scalingMethod': "StandardScaler", 
        'QuantileTransformer': "True", 
        'Normalize': "True",
        'rebalance' : "True",
        'rebalance_type': "SMOTE",
        'FeatureSelection': "True",
        "method": 'SelectKBest-f_classif',
        'N_features': 40, 
        'strategy': "random", 
        'itr': 50,
        "CV": 10,
        "n_repeats": 1,
        "min_postives": 10,
    }

options_default = { # test set options dict
        'oneHotEncode' : "True", 
        'Impute': "True", 
        'cutMissingRows' : "True",
        "cut threshold": 0.60,
        "inf": 'replace with null',
        'outliers': "log",
        'outliers_N': 50000, 
        'Scaling': "True", 
        'scalingMethod': "StandardScaler", 
        'QuantileTransformer': "True", 
        'Normalize': "True",
        'rebalance' : "True",
        'rebalance_type': "SMOTE",
        'FeatureSelection': "True",
        "method": 'SelectKBest-f_classif',
        'N_features': 40, 
        'strategy': "random", 
        'itr': 50,
        "CV": 10,
        "n_repeats": 1,
        "min_postives": 10,
        "test_size": 0.2,
        'cutoff_index': 'youden'
    }

def sanitize_filename(filename):
    """Remove or replace invalid characters from filenames."""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    return filename

def shorten_component(text, max_len=50):
    """Shortens a component to max_len safely, using hash suffix if needed."""
    if len(text) <= max_len:
        return text
    hash_suffix = hashlib.md5(text.encode()).hexdigest()[:6]
    return text[:max_len - 7] + "_" + hash_suffix

# Python code to remove whitespace
def remove(string):
    return string.replace(" ", "")

def apply_coding(value, bins, code):
    for i, b in enumerate(bins):
        if value in bins[i]:
            return code[i]
        
def varname(p):
    for line in inspect.getframeinfo(inspect.currentframe().f_back)[3]:
        m = re.search(r'\bvarname\s*\(\s*([A-Za-z_][A-Za-z0-9_]*)\s*\)', line)
        if m:
            return m.group(1)

# Function to get machine specs
def get_machine_specs():
    cpu_info = psutil.cpu_percent()
    mem_info = psutil.virtual_memory().total / (1024 ** 3)  # Convert bytes to GB
    return cpu_info, mem_info

def dict_to_excel(data, file_path=""):
    # Convert dictionary to DataFrame
    df = pd.DataFrame(data).T

    # Flatten lists in the DataFrame (if any)
    for column in df.columns:
        df[column] = df[column].apply(lambda x: str(x) if isinstance(x, list) else x)

    # Write DataFrame to Excel
    df.to_excel(file_path)
    
    return df

def wrap_text_excel(file_path):
    # Load the workbook and the sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Set text wrapping for all cells and adjust column width based on the longest cell in each column
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter

        for cell in col:
            # Set wrap_text to True for all cells
            cell.alignment = Alignment(wrap_text=True)
            
            # Check the length of the cell value (if it's not None)
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        # Adjust the column width based on the longest value, considering an approximation factor
        adjusted_width = (max_length + 2) * 1.1  # Add padding and adjust width factor
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the workbook
    wb.save(file_path)

def expand_cell_excel(file_path):
    # Load the workbook and the sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Adjust the width of the columns based on the max length in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(file_path)

def grid_excel(file_path):
    # Load the workbook and the sheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define border style
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))
    
    # Apply borders to all cells to create a grid
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            
    # Save the workbook
    wb.save(file_path)

# Function to fill cells with color if their values are in the target list
def fill_cells_with_color(file_path, target_values, color):
    # Load the workbook and select the specified sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Define the fill pattern with the specified color
    fill_pattern = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Iterate over all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # If the cell value is in the target list, apply the fill pattern
            if cell.value in target_values:
                cell.fill = fill_pattern

    # Save the modified workbook
    workbook.save(file_path)

def parse_exp_csv(df, filename, project_name, unique_value_threshold=10): # Define the threshold for considering discrete numeric columns as categorical
    print(filename[-4:])
    # upload index file
    pathname = project_name + '/' + filename[:-4] + "_index.csv"
    df_exp = upload_data(pathname)
    
    # inputs and outputs for a row
    # return index
    input_cols = list(df_exp.columns[(df_exp == 1).all()])
    outputs_cols = list(df_exp.columns[(df_exp == 2).all()])[0]
    
    input_df = df.loc[:, input_cols]

    # Select columns of object data type (categorical columns)
    categorical_cols = input_df.select_dtypes('object').columns.tolist()

    # Identify discrete numeric columns as categorical if they have fewer unique values than the threshold
    discrete_numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if len(input_df[col].unique()) < unique_value_threshold]
    categorical_cols.extend(discrete_numeric_cols)

    # Exclude the identified discrete numeric columns from numeric_cols
    numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if col not in discrete_numeric_cols]
    
    return input_cols, outputs_cols, categorical_cols, numeric_cols

def parse_exp(df, filename, project_name, unique_value_threshold=10): # Define the threshold for considering discrete numeric columns as categorical
    file_type = filename[-4:]
    print("File Type:", file_type)
    # upload index file
    if file_type == "xlsx":
        pathname = project_name + '/' + filename[:-5] + "_index.xlsx"
    else:
        pathname = project_name + '/' + filename[:-4] + "_index.csv"
    df_exp = upload_data(pathname)
    
    # inputs and outputs for a row
    # return index
    input_cols = list(df_exp.columns[(df_exp == 1).all()])
    outputs_cols = list(df_exp.columns[(df_exp == 2).all()])[0]
    
    input_df = df.loc[:, input_cols]

    # Select columns of object data type (categorical columns)
    categorical_cols = input_df.select_dtypes('object').columns.tolist()

    # Identify discrete numeric columns as categorical if they have fewer unique values than the threshold
    discrete_numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if len(input_df[col].unique()) < unique_value_threshold]
    categorical_cols.extend(discrete_numeric_cols)

    # Exclude the identified discrete numeric columns from numeric_cols
    numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if col not in discrete_numeric_cols]
    
    return input_cols, outputs_cols, categorical_cols, numeric_cols

def parse_exp_multi_outcomes(df, df_index, unique_value_threshold=10): # Define the threshold for considering discrete numeric columns as categorical
    
    # inputs and outputs for a row
    # return index
    input_cols = list(df_index.columns[(df_index == 1).all()])
    outputs_cols = list(df_index.columns[(df_index == 2).all()])
    
    input_df = df.loc[:, input_cols]

    # Select columns of object data type (categorical columns)
    categorical_cols = input_df.select_dtypes('object').columns.tolist()

    # Identify discrete numeric columns as categorical if they have fewer unique values than the threshold
    discrete_numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if len(input_df[col].unique()) < unique_value_threshold]
    categorical_cols.extend(discrete_numeric_cols)

    # Exclude the identified discrete numeric columns from numeric_cols
    numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if col not in discrete_numeric_cols]
    
    return input_cols, outputs_cols, categorical_cols, numeric_cols

def save_dictionary(dict_file, file_type=".json"):
    # get name of the project as part of file name
    project_name = list(dict_file.keys())[0]
    
    if file_type == ".json":
        # Convert Python to JSON  
        json_data = json.dumps(dict_file, indent = 4)

        # write the JSON string to a file
        with open(project_name + '.json', 'w') as f:
            f.write(json_data)
            
    elif file_type == ".txt":
        # Path to save the file
        file_path = project_name + '.txt'
        
        # Convert the dictionary to a string representation
        dict_str = str(dict_file)

        # Save the string representation to a text file
        with open(file_path, 'w') as file:
            file.write(dict_str)
            
    # Return the name of the file
    filename = project_name + file_type
    print(filename)
    return filename

def generate_configuration_file_old(num_exp, project_name, test_set, exp_name, algorithms, filename, problem_type, cross_validation, options, param_vals):
    configuration_dic = {}
    configuration_dic[project_name] = {}
    
    for i in range(num_exp):
        #st.write(i)
        experiment = {}
        experiment['algorithm'] = algorithms[i]
        experiment['test_set'] = test_set[i]
        experiment['filename'] = filename[i]
        experiment['problem_type'] = problem_type[i]
        experiment['cross_validation'] = cross_validation[i]
        experiment['options'] = options[i]
        experiment['param_vals'] = param_vals[i]
        
        configuration_dic[project_name][exp_name] = experiment
    
    print(configuration_dic)
    
    return configuration_dic
    

# find average values acorss all arrays, even if they're not in the same shape
def average_values_across_lists(arrays):
    # Find the maximum length among all arrays
    max_length = max(len(arr) for arr in arrays)
    
    # Pad shorter arrays with np.nan values to match the length of the longest array
    #padded_arrays = [arr + [np.nan] * (max_length - len(arr)) for arr in arrays]
    padded_arrays = [np.concatenate((arr, [np.nan] * (max_length - len(arr)))) if len(arr) < max_length else arr for arr in arrays]
    

    # Convert to NumPy array for easy calculation
    arr = np.array(padded_arrays)

    # Calculate the mean along axis 0 (across lists)
    average_values = np.nanmean(arr, axis=0)

    return average_values

# find average values acorss all 2D arrays, even if they're not in the same shape
def average_values_across_arrays(arrays):
    # Extract the 2D NumPy array from the input list
    matrix = arrays[0]

    # Find the maximum dimensions along both axes
    max_rows = len(matrix)
    max_cols = max(len(row) for row in matrix)

    # Pad shorter rows and columns with np.nan values to match the maximum dimensions
    padded_matrix = [
        np.concatenate((row, [np.nan] * (max_cols - len(row)))) if len(row) < max_cols else row
        for row in matrix
    ]
    padded_matrix.extend([[np.nan] * max_cols] * (max_rows - len(matrix)))

    # Convert to NumPy array for easy calculation
    arr = np.array(padded_matrix)

    # Calculate the mean along both axes
    average_values = np.nanmean(arr, axis=(0, 1))

    return average_values

def resize_arrays_to_smallest(arrays):
    # Get the minimum number of rows and columns among all arrays
    min_rows = min(array.shape[0] for array in arrays)
    min_cols = min(array.shape[1] for array in arrays)

    # Resize each array to match the minimum size
    resized_arrays = []
    for array in arrays:
        # Truncate rows if necessary
        if array.shape[0] > min_rows:
            array = array[:min_rows, :]
        
        # Truncate columns if necessary
        if array.shape[1] > min_cols:
            array = array[:, :min_cols]

        resized_arrays.append(array)

    return resized_arrays

# By default, we should cut rows that have missing series of values (cutMissing), impute the data (Impute), and remove very large data values (removeBig)
def preprocess(df, input_cols, label_cols, numeric_cols, categorical_cols, cutMissingRows='True', threshold=0.75, oneHotEncode='True', inf='replace with null', outliers='None', N=20000, QuantileTransformer='False'):
    if oneHotEncode == 'True':
        print("oneHotEncode")
        # One Hot Encode catagorical variables
        from sklearn.preprocessing import OneHotEncoder
        encoder = OneHotEncoder(sparse_output=False, handle_unknown='ignore')
        encoder.fit(df[categorical_cols])
        encoded_cols = list(encoder.get_feature_names_out(categorical_cols))
        df[encoded_cols] = encoder.transform(df[categorical_cols])
        input_cols = numeric_cols + encoded_cols
    
    '''
    if Impute == True:
        print("Impute")
        # Impute the remaining missing numeric data
        from sklearn.impute import SimpleImputer
        imputer = SimpleImputer(strategy = 'mean')
        imputer.fit(df[numeric_cols])
        df[numeric_cols] = imputer.transform(df[numeric_cols])
    '''
        
    if cutMissingRows == 'True':
        print("cutMissingRows")
        # Drop rows with missing values
        # computing number of columns
        cols = len(df.axes[1])
        print("Cuttoff", int(threshold * cols))
        df = df.dropna(thresh=int(threshold * cols))

    #print(df)

    if inf == 'replace with null':
        print("replace with null")
        # Replace all inf values with null
        df.replace([np.inf, -np.inf], np.nan, inplace=True)
    elif inf == 'replace with zero':
        print("replace with zero")
        # Replace all inf values with null
        df.replace([np.inf, -np.inf], 0, inplace=True)

    #print(df)

    if outliers == 'remove rows':
        print("remove rows")
        # Remove rows that have a value greater than N for any column. Default N is 20000
        for column  in df[numeric_cols]:
            df = df.drop(df.index[df[column] > N])
    elif outliers == 'log':
        print("log")
        # Log values that are greater than N for any column. Default N is 20000
        df[numeric_cols].apply(lambda x: np.where(x > N, np.log(x), x))

    #print(df)

    '''
    if Scaling == True:
        print("Scaling")
        # Scaling the input features for a chosen method. Default is MinMaxScaler.
        if scalingMethod == 'MinMaxScaler': 
            print("MinMaxScaler")
            # This estimator scales and translates each feature individually such that it is in the given range on the training set, e.g. between zero and one.
            from sklearn.preprocessing import MinMaxScaler
            scaler = MinMaxScaler()
        elif scalingMethod == 'RobustScaler':
            print("RobustScaler")
            # This Scaler removes the median and scales the data according to the quantile range (defaults to IQR: Interquartile Range). The IQR is the range between the 1st quartile (25th quantile) and the 3rd quartile (75th quantile).
            from sklearn.preprocessing import RobustScaler
            scaler = RobustScaler()
        elif scalingMethod == 'MaxAbsScaler':
            print("MaxAbsScaler")
            # Scale each feature by its maximum absolute value.
            from sklearn.preprocessing import MaxAbsScaler
            scaler = MaxAbsScaler()
        elif scalingMethod == 'StandardScaler':
            print("StandardScaler")
            # Standardize features by removing the mean and scaling to unit variance.
            from sklearn.preprocessing import StandardScaler
            scaler = StandardScaler()
        scaler.fit(df[numeric_cols])
        df[numeric_cols] = scaler.transform(df[numeric_cols])
    '''
        
    if QuantileTransformer == 'True':  
        print("QuantileTransformer")
        # transforms the features to follow a uniform or a normal distribution.
        from sklearn.preprocessing import QuantileTransformer
        qt = QuantileTransformer(output_distribution='normal').fit(df[numeric_cols])
        df[numeric_cols] = qt.transform(df[numeric_cols])
    '''   
    if Normalize == 'True':  
        print("Normalize")
        # Normalize the data
        from sklearn.preprocessing import Normalizer
        normalizer = Normalizer().fit(df[numeric_cols])
        df[numeric_cols] = normalizer.transform(df[numeric_cols])
    '''    
    
    return df, input_cols


# Version of of the preprocess function that's used for senarios of seperate train and test sets
def preprocess_train_test(df, df_test, input_cols, label_cols, numeric_cols, categorical_cols, cutMissingRows='True', threshold=0.75, oneHotEncode='True', inf='replace with null', outliers='None', N=20000, QuantileTransformer='False'):
    if oneHotEncode == 'True':
        print("oneHotEncode")
        # One Hot Encode catagorical variables
        from sklearn.preprocessing import OneHotEncoder
        encoder = OneHotEncoder(sparse_output=False, handle_unknown='ignore')
        encoder.fit(df[categorical_cols])
        encoded_cols = list(encoder.get_feature_names_out(categorical_cols))
        df[encoded_cols] = encoder.transform(df[categorical_cols])
        df_test[encoded_cols] = encoder.transform(df_test[categorical_cols])
        input_cols = numeric_cols + encoded_cols
        
    if cutMissingRows == 'True':
        print("cutMissingRows")
        print("Inital Test Data Size", len(df_test))
        # Drop rows with missing values
        # computing number of columns
        cols = len(df_test.axes[1])
        print("Number of Inital Columns", cols)
        print("Cuttoff", int(threshold * cols))
        df = df.dropna(thresh=int(threshold * cols))
        df_test = df_test.dropna(thresh=int(threshold * cols))
        print("Number of Final Columns", len(df_test.axes[1]))
        print("Final Test Data Size", len(df_test))
    
    if inf == 'replace with null':
        print("replace with null")
        # Replace all inf values with null
        df.replace([np.inf, -np.inf], np.nan, inplace=True)
        df_test.replace([np.inf, -np.inf], np.nan, inplace=True)
    elif inf == 'replace with zero':
        print("replace with zero")
        # Replace all inf values with null
        df.replace([np.inf, -np.inf], 0, inplace=True)
        df_test.replace([np.inf, -np.inf], 0, inplace=True)
    #print(df)

    if outliers == 'remove rows':
        print("remove rows")
        # Remove rows that have a value greater than N for any column. Default N is 20000
        for column  in df[numeric_cols]:
            df = df.drop(df.index[df[column] > N])
            df_test = df_test.drop(df_test.index[df_test[column] > N])
    elif outliers == 'log':
        print("log")
        # Log values that are greater than N for any column. Default N is 20000
        df[numeric_cols].apply(lambda x: np.where(x > N, np.log(x), x))
        df_test[numeric_cols].apply(lambda x: np.where(x > N, np.log(x), x))

    #print(df)
        
    if QuantileTransformer == 'True':  
        print("QuantileTransformer")
        # transforms the features to follow a uniform or a normal distribution.
        from sklearn.preprocessing import QuantileTransformer
        qt = QuantileTransformer(output_distribution='normal').fit(df[numeric_cols])
        df[numeric_cols] = qt.transform(df[numeric_cols])
        df_test[numeric_cols] = qt.transform(df_test[numeric_cols])
    '''    
    if Normalize == 'True':  
        print("Normalize")
        # Normalize the data
        from sklearn.preprocessing import Normalizer
        normalizer = Normalizer().fit(df[numeric_cols])
        df[numeric_cols] = normalizer.transform(df[numeric_cols])
        df_test[numeric_cols] = normalizer.transform(df_test[numeric_cols])
    '''
    return df, df_test, input_cols

# Split data into input and output sets
def split(df, input_cols, label_cols):
    input_df = df[input_cols].copy()
    output_df = df[label_cols].copy()
    
    return input_df, output_df


def scaling(df, input_cols, label_cols, numeric_cols, categorical_cols, scalingMethod='MinMaxScaler'):
        print("Scaling")
        # Scaling the input features for a chosen method. Default is MinMaxScaler.
        if scalingMethod == 'MinMaxScaler': 
            print("MinMaxScaler")
            # This estimator scales and translates each feature individually such that it is in the given range on the training set, e.g. between zero and one.
            from sklearn.preprocessing import MinMaxScaler
            scaler = MinMaxScaler()
        elif scalingMethod == 'RobustScaler':
            print("RobustScaler")
            # This Scaler removes the median and scales the data according to the quantile range (defaults to IQR: Interquartile Range). The IQR is the range between the 1st quartile (25th quantile) and the 3rd quartile (75th quantile).
            from sklearn.preprocessing import RobustScaler
            scaler = RobustScaler()
        elif scalingMethod == 'MaxAbsScaler':
            print("MaxAbsScaler")
            # Scale each feature by its maximum absolute value.
            from sklearn.preprocessing import MaxAbsScaler
            scaler = MaxAbsScaler()
        elif scalingMethod == 'StandardScaler':
            print("StandardScaler")
            # Standardize features by removing the mean and scaling to unit variance.
            from sklearn.preprocessing import StandardScaler
            scaler = StandardScaler()
        scaler.fit(df[numeric_cols])
        df[numeric_cols] = scaler.transform(df[numeric_cols])
        
        return df, scaler

#rebalance the imbalanced data with a chosen label
def rebalance(input_df, label_df, type='RandomUnderSampler'):
    if type == 'RandomUnderSampler':
        # random undersampling reduces the number of majority class randomly down to the desired ratio against the minority class.
        from imblearn.under_sampling import RandomUnderSampler
        rebalance = RandomUnderSampler()
        input_df2, label_df2 = rebalance.fit_resample(input_df, label_df)
    elif type == 'RandomOverSampler':
        # Naive random over-sampling.
        from imblearn.over_sampling import RandomOverSampler
        rebalance = RandomOverSampler()
        input_df2, label_df2 = rebalance.fit_resample(input_df, label_df)
    elif type == 'SMOTE':
        # SMOTE is a technique to up-sample the minority classes while avoiding overfitting.
        from imblearn.over_sampling import SMOTE
        rebalance = SMOTE()
        input_df2, label_df2 = rebalance.fit_resample(input_df, label_df)    
    elif type == 'ADASYN':
        # Adaptive Synthetic (ADASYN) algorithm. This method is similar to SMOTE but it generates different number of samples depending on an estimate of the local distribution of the class to be oversampled.
        from imblearn.over_sampling import ADASYN
        rebalance = ADASYN()
        input_df2, label_df2 = rebalance.fit_resample(input_df, label_df)                   
    else:
        print("Cannot do")
        input_df2, label_df2 = input_df, label_df
        
    return input_df2, label_df2  


def feature_selection(input_df, output_df, method="MRMR", type='f_classif', N=20, per=10):
    if method=="MRMR":
        selected_features = mrmr_classif(X=input_df, y=output_df, K=N)
        input_df_new = input_df[selected_features]
    elif method=="SelectKBest":
        if type=='f_classif':
            best_features =  SelectKBest(f_classif, k=N)
            best_features.fit(input_df, output_df)
        elif type=='chi2':
            best_features =  SelectKBest(chi2, k=N)
            best_features.fit(input_df, output_df)
        # Get columns to keep and create new dataframe with those only
        cols_idxs = best_features.get_support(indices=True)
        input_df_new = input_df.iloc[:,cols_idxs]
    elif method=="SelectPercentile": # Select features according to a percentile of the highest scores.
        if type=='f_classif':
            best_features =  SelectPercentile(f_classif, percentile=per)
            best_features.fit(input_df, output_df)
        elif type=='chi2':
            best_features =  SelectPercentile(chi2, percentile=per)
            best_features.fit(input_df, output_df)
        # Get columns to keep and create new dataframe with those only
        cols_idxs = best_features.get_support(indices=True)
        input_df_new = input_df.iloc[:,cols_idxs]
    elif method=="VarianceThreshold": # Feature selector that removes all low-variance features.
        best_features = VarianceThreshold()
        best_features.fit(input_df)
        cols_idxs = best_features.get_support(indices=True)
        input_df_new = input_df.iloc[:,cols_idxs]
    
    return input_df_new, list(input_df_new.columns)


# calibrate the model
# Probability calibration is the process of calibrating an ML model to return the true likelihood of an event.
def calibrate(estimator, X_train, y_train, cal_method='sigmoid', cv=None):
    cal_estimator = CalibratedClassifierCV(base_estimator=estimator, method=cal_method, cv=cv)
    cal_estimator.fit(X_train, y_train)  
    return cal_estimator


def train_tune(estimator, param_vals, X_train, y_train, strategy='random', itr=20, cv=None):
    if strategy == 'random':
        random_df = RandomizedSearchCV(estimator, param_distributions=param_vals, cv=cv,
                              n_iter=itr, random_state=256, n_jobs=-1)
        random_df.fit(X_train, y_train)
        best_model = random_df.best_estimator_

    elif strategy == 'bayesian':
        np.int = int # to avoid int decaprated error
        bayes_df = BayesSearchCV(estimator, param_vals, cv=cv,
                              n_iter=itr, random_state=256, n_jobs=-1)
        bayes_df.fit(X_train, y_train)
        best_model = bayes_df.best_estimator_

    elif strategy == 'grid':
        grid_df = GridSearchCV(estimator, param_vals, n_jobs=-1, cv=cv, return_train_score=True)
        grid_df.fit(X_train, y_train)
        best_model = grid_df.best_estimator_

    elif strategy == 'ray':
        grid_df = TuneSearchCV(estimator=estimator,
                               param_distributions=param_vals,
                               n_trials=itr,
                               n_jobs=-1,
                               verbose=2)

        grid_df.fit(X_train, y_train)
        best_model = grid_df.best_estimator_ 

    else:
        random_df = RandomizedSearchCV(estimator, param_distributions=param_vals, cv=cv,
                              n_iter=itr, random_state=256, n_jobs=-1)
        random_df.fit(X_train, y_train)
        best_model = random_df.best_estimator_
        
    return best_model

def train_tune_cv(estimator, param_vals, X_train, y_train, strategy='random', itr=20, cv=None, scoring='roc_auc'):
    if strategy == 'random':
        random_df = RandomizedSearchCV(estimator, param_distributions=param_vals, cv=cv, scoring=scoring,
                              n_iter=itr, random_state=256, n_jobs=-1, return_train_score=True)
        random_df.fit(X_train, y_train)
        return random_df
    
    elif strategy == 'bayesian':
        np.int = int # to avoid int decaprated error
        bayes_df = BayesSearchCV(estimator, param_vals, cv=cv, scoring=scoring,
                              n_iter=itr, random_state=256, n_jobs=-1, return_train_score=True)
        bayes_df.fit(X_train, y_train)
        return bayes_df
    
    elif strategy == 'grid':
        grid_df = GridSearchCV(estimator, param_vals, n_jobs=-1, cv=cv, scoring=scoring, return_train_score=True)
        grid_df.fit(X_train, y_train)
        return grid_df

    elif strategy == 'ray':
        grid_df = TuneSearchCV(estimator=estimator,
                               param_distributions=param_vals,
                               n_trials=itr,
                               n_jobs=-1,
                               verbose=2)

        grid_df.fit(X_train, y_train)
        return grid_df
    
    else:
        random_df = RandomizedSearchCV(estimator, param_distributions=param_vals, cv=cv, scoring=scoring,
                              n_iter=itr, random_state=256, n_jobs=-1, return_train_score=True)
        random_df.fit(X_train, y_train)
        return random_df


def upload_data(filename):
    print(filename)
    file_type = magic.from_file(filename)
    print(file_type)
    if file_type[:10] == 'ASCII text': # if file type if a .csv
        print('.csv')
        raw_df = pd.read_csv(filename)
    elif file_type == 'Microsoft Excel 2007+': # if file type is a .xlsx
        print(".xlsx")
        raw_df = pd.read_excel(filename)
    else:
        print('.csv')
        raw_df = pd.read_csv(filename)
    return raw_df.copy()

def load_data(filename, data):
    if filename.endswith('.xlsx'):
        df = pd.read_excel(data)
    elif filename.endswith('.csv'):
        df = pd.read_csv(data)
    else:
        df = pd.read_excel(data)

    return df

def save_data(filename, data, save_path):
    if filename.endswith('.xlsx'):
        data.to_excel(save_path)
    elif filename.endswith('.csv'):
        data.to_csv(save_path)
    else:
        data.to_csv(save_path)




def open_configuration_file(filename):
    print(filename[-5:])
    if filename[-5:] == '.json': # if file is a JSON file
        # Opening JSON file
        f = open(filename)

        # returns JSON object as 
        # a dictionary
        all_experiments = json.load(f)
        
    elif filename[-4:] == '.txt':
        # Load the saved dictionary from the file
        with open(filename, 'r') as file:
            loaded_data_str = file.read()
            all_experiments = eval(loaded_data_str)  # Convert the string back to a dictionary using eval

            print("Dictionary loaded successfully:")
            #print(loaded_data)
        
    return all_experiments


def data_prep(df, input_cols, label_cols, numeric_cols, categorical_cols, options):
    #Preprocess data
    df_new, input_cols = preprocess(df, input_cols, label_cols, numeric_cols, categorical_cols, oneHotEncode=options['oneHotEncode'],
                    cutMissingRows=options['cutMissingRows'], threshold=options['cut threshold'], inf=options['inf'],
                    outliers=options['outliers'], N=options['outliers_N'],
                        QuantileTransformer=options['QuantileTransformer'])

    return df_new, input_cols, label_cols

# Version of of the data_prep function that's used for senarios of seperate train and test sets
def data_prep_train_set(df, def_test, input_cols, label_cols, numeric_cols, categorical_cols, options):
    #Preprocess data
    df_new, df_new_test, input_cols = preprocess_train_test(df, def_test, input_cols, label_cols, numeric_cols, categorical_cols, oneHotEncode=options['oneHotEncode'],
                    cutMissingRows=options['cutMissingRows'], threshold=options['cut threshold'], inf=options['inf'],
                    outliers=options['outliers'], N=options['outliers_N'],
                        QuantileTransformer=options['QuantileTransformer'])

    return df_new, df_new_test, input_cols, label_cols

# get the correlation matrix
def correlation_matrix(df, numeric_cols, label_cols, algorithm_folder):
    # Compute the correlation matrix
    ness_cols = numeric_cols.copy()
    ness_cols.append(label_cols)
    correlation_matrix = df[ness_cols].corr()

    # Set up the figure and axes
    plt.figure(figsize=(10, 8))

    # Create the heatmap using Seaborn
    sns.heatmap(correlation_matrix, annot=False, cmap='coolwarm', fmt=".2f")
    
    # Adjust font size dynamically based on the size of the heatmap cells
    n_rows, n_cols = correlation_matrix.shape
    cell_size = 1.0 / max(n_rows, n_cols)
    font_size = 150 * cell_size  # Adjust multiplier as needed

    # Annotate values inside the matrix with count and percentage
    for i in range(len(correlation_matrix)):
        for j in range(len(correlation_matrix)):
            plt.text(j + 0.5, i + 0.5, f'{correlation_matrix.iloc[i, j]:.2f}', horizontalalignment='center', verticalalignment='center', color='black', fontsize=font_size)

    # Set plot title
    plt.title('Correlation Heatmap')
    
    # Save the figure
    algorithm_folder = algorithm_folder
    # Save the combined figure as a .png file
    save_path = os.path.join(algorithm_folder, 'correlation_matrix.png')
    plt.savefig(save_path)
    plt.show()

# get the top 20 varaibles correlated with the final output label to predict
def top_correlation_variables(df, numeric_cols, label_cols, algorithm_folder):
    # Compute the correlation matrix
    ness_cols = numeric_cols.copy()
    ness_cols.append(label_cols)
    correlation_matrix = df[ness_cols].corr()
    
    # Calculate the correlation of 'A' with other columns
    corr_with_label = df[numeric_cols].corrwith(df[label_cols])
    
    # Get the top 20 variables corr. with the output label
    top_corr_variables = corr_with_label.abs().sort_values(ascending=False)[:21]

    # Print the top 20 variables corr. with column label
    print("Top 20 Correlated Variables with " + label_cols + " (By absolute values) :")
    print(top_corr_variables)
    
    # Write to text file
    algorithm_folder = algorithm_folder
    with open(os.path.join(algorithm_folder, "top_correlated_variables.txt"), "w") as file:
        file.write("Top 20 Correlated Variables with '{}' (By absolute values) :\n".format(label_cols))
        for variable, correlation in top_corr_variables.items():
            file.write("{} : {}\n".format(variable, correlation))

#dummy model with prediction of all 0's
def all_zero(inputs, output):
    answer = output.unique()[0]
    return np.full(len(inputs), 0)

#dummy model with prediction of all 1's
def all_ones(inputs, output):
    answer = output.unique()[0]
    return np.full(len(inputs), 1)

def load_image(image_path):
    try:
        # Open the image file
        img = Image.open(image_path)
            
        # Display basic information about the image
        print("Image Format:", img.format)
        print("Image Mode:", img.mode)
        print("Image Size:", img.size)
        
        # Show the image
        #img.show()
        
        # Display the image directly within the notebook
        display(img)
        
    except FileNotFoundError:
        print("File not found. Please check the file path.")

def merge_images(image_paths, output_path):
    images = [Image.open(path) for path in image_paths]
    
    # Determine common height if not provided
    common_height = max(img.size[1] for img in images)

    # Resize images to the common height
    resized_images = [img.resize((int(img.size[0] * common_height / img.size[1]), common_height)) for img in images]

    # Get total width of the merged image
    total_width = sum(img.size[0] for img in resized_images)

    # Create a new blank image with dimensions large enough to accommodate all images
    merged_image = Image.new("RGBA", (total_width, common_height))

    # Paste each resized image onto the blank image
    x_offset = 0
    for img in resized_images:
        merged_image.paste(img, (x_offset, 0))
        x_offset += img.size[0]

    # Save the merged image
    merged_image.save(output_path)
    
    merged_image.show()

def merge_images_grid(image_paths, output_path, columns=3):
    images = [Image.open(path) for path in image_paths]
    
    # Determine common height if not provided
    common_height = max(img.size[1] for img in images)

    # Resize images to the common height
    resized_images = [img.resize((int(img.size[0] * common_height / img.size[1]), common_height)) for img in images]
    
    # Calculate number of rows needed
    rows = (len(resized_images) + columns - 1) // columns

    # Get total dimensions of the merged image
    total_width = max(img.size[0] for img in resized_images) * columns
    total_height = common_height * rows

    # Create a new blank image with dimensions large enough to accommodate all images
    merged_image = Image.new("RGBA", (total_width, total_height))

    # Paste each resized image onto the blank image
    x_offset = 0
    y_offset = 0
    for img in resized_images:
        merged_image.paste(img, (x_offset, y_offset))
        x_offset += img.size[0]
        if x_offset >= total_width:
            x_offset = 0
            y_offset += common_height

    # Save the merged image
    merged_image.save(output_path)
    
    merged_image.show()


def check_same_feature_set(df1, df2):
    # Get the column names of both DataFrames
    features_df1 = set(df1.columns)
    features_df2 = set(df2.columns)
    
    # Check if the sets of column names are equal
    if features_df1 == features_df2:
        print("Both DataFrames have the same feature set.")
        return True
    else:
        print("The feature sets of the DataFrames are not the same.")
        different_columns_df1 = features_df1 - features_df2
        different_columns_df2 = features_df2 - features_df1
        
        if different_columns_df1:
            print("Columns in df1 but not in df2:", different_columns_df1)
        if different_columns_df2:
            print("Columns in df2 but not in df1:", different_columns_df2)
        
        return False
    
def gen_df_index_csv(filename:str, nexp:int=1, project_name=None):
    '''

    Parameters
    ----------
    filename : string
        data file path and name.
    nexp : int, optional
        Number of experiments. The default is 1.

    Returns
    -------
    index file path and name.

    '''
    # Check if the file exists
    if not os.path.exists(filename):
        raise FileNotFoundError(f"The file '{filename}' does not exist.")
    
    # Load the CSV file
    df_original = pd.read_csv(filename)
    
    # Drop 'Unnamed: 0' column if it exists
    if 'Unnamed: 0' in df_original.columns:
        df_original.drop(columns=['Unnamed: 0'], axis=1, inplace=True)
    
    indexFileName = os.path.splitext(filename)[0] + '_index.csv'
    
    #algorithm_folder = os.path.join(project_name)
    #os.makedirs(algorithm_folder, exist_ok=True)  # Create folder for algorithm results
    
    with open(os.path.join(project_name, indexFileName), 'w') as f:
        # Write header to the index file
        f.write(','.join([''] + list(df_original.columns)) + '\n')
        
        # Write experiment rows to the index file
        for i in range(nexp):
            f.write(','.join(['Exp' + str(i+1)] + ['0']*(len(df_original.columns))) + '\n')
    
    return indexFileName

def gen_df_index_excel(filename:str, nexp:int=1, sheets:list=None, project_name=None):
    '''

    Parameters
    ----------
    filename : string
        data file path and name.
    nexp : int, optional
        Number of experiments. The default is 1.
    sheets : list, optional
        list of sheets to be used. The default is None, which will use all sheets.

    Returns
    -------
    index file path and name.

    '''
    
    book = load_workbook(filename=filename, read_only=True, data_only=True)
    indexFileName = filename[:-5] + '_index.xlsx'
    #writer = pd.ExcelWriter(indexFileName, engine='xlsxwriter')
    
    #algorithm_folder = os.path.join(project_name)
    #os.makedirs(algorithm_folder, exist_ok=True)  # Create folder for algorithm results
    
    with pd.ExcelWriter(os.path.join(project_name, indexFileName), engine='xlsxwriter') as writer:
        if sheets == None:
            sheets = range(len(book.worksheets))
        for idx in sheets:
            first_sheet = book.worksheets[idx]
            rows_generator = first_sheet.values

            header_row = next(rows_generator)
            ncols = len(header_row)
            row_name = ['Exp'+str(x+1) for x in range(nexp)]
            df_exp = pd.DataFrame(np.zeros((nexp, ncols)), columns=header_row, index=row_name) 

            df_exp.to_excel(writer, sheet_name=book.sheetnames[idx])
    
    writer.close()  # Close the file handle
    
    return indexFileName
    
def gen_idx(filename, project_name):
    print(filename)
    file_type = magic.from_file(filename)
    print(file_type)
    if file_type[:10] == 'ASCII text': # if file type if a .csv
        gen_df_index_csv(filename=filename, project_name=project_name)
    elif file_type == 'Microsoft Excel 2007+': # if file type is a .xlsx
        gen_df_index_excel(filename=filename, project_name=project_name)
    else:
        gen_df_index_csv(filename=filename, project_name=project_name)

    
def generate_all_idx_files(all_experiments, project_name):
    experiment_names = list(all_experiments[project_name].keys())
    print("List of experiments: ", experiment_names)
    
    for experiment in all_experiments[project_name]:
        experiment_name = experiment
        print(experiment_name)
        set_up = all_experiments[project_name][experiment_name]
        
        train_set = set_up['train_set']
        df = upload_data(train_set)
        
        gen_idx(train_set, project_name)
            
        print("_____________________________________________________________________________________________________")

def setup_multioutcome_binary(set_up, experiment_name, project_folder):
    algorithm = set_up['algorithm']
    st.write(algorithm)
    
    options = set_up['options']
    
    param_vals = set_up['param_vals']
    
    # Write to text file
    algorithm_folder = os.path.join(project_folder, experiment_name)
    os.makedirs(algorithm_folder, exist_ok=True)  # Create folder for algorithm results
    with open(os.path.join(algorithm_folder, "parameter_setup.txt"), "w", encoding="utf-8") as file:
        file.write('Experiment Name: %s\n' % experiment_name)
        file.write('ML Algorithm: %s\n' % algorithm)
        for key, value in options.items(): 
            file.write('%s:%s\n' % (key, value))
    file.close()

    return options, algorithm, param_vals


def refine_binary_outcomes(df, label_cols):
    for label_col in label_cols:
        if label_col in df.columns and not(df[label_col].isin([0, 1]).all()):
            df[label_col] = df[label_col].fillna(df[label_col].mode()[0])
            
            first_label = df[label_col].unique()[0]
            second_label = df[label_col].unique()[1]
    
            if first_label in df[label_col].tolist() and second_label in df[label_col].tolist():
                df[label_col] = df[label_col].map({first_label: 0, second_label: 1}).astype(int)

    #st.write(df)
    return df


# function to generate and return a joblib file storing all the models created in the experiment
def generate_joblib_model(directory_path):
    # Create directory to store all the ML content
    model_dic = {}
    input_cols_dic = {}

    # Get the name of the directory
    directory_name = os.path.basename(directory_path)

    # Get the list of contents in the directory
    directory_contents = os.listdir(directory_path)

    print(f"Directory Name: {directory_name}")

    for item_name in directory_contents:
        print(item_name)
        item_path = os.path.join(directory_path, item_name)

        if os.path.isdir(item_path):
            #model_dic[item_name] = {}

            print("Directory Contents:")
            #print(item_path)
            # Get the list of contents in the directory
            item_contents = os.listdir(item_path)

            print("   Item Contents:")
            for content_name in item_contents:
                content_path = os.path.join(directory_path, item_name, content_name)
                if os.path.isdir(content_path) and content_path.find("(results)") == -1:

                    print("  ",content_path)
                    algorithm = content_name
                    print("  ",algorithm)
                    model_dic[algorithm] = {}
                    
                    # get the input columns
                    try:
                        input_columns = joblib.load(os.path.join(directory_path, item_name, 'input_columns.joblib'))
                        print(f'Input Columns for {item_path}: {input_columns}')
                        input_cols_dic[algorithm] = input_columns
                    except:
                        print("      No input columns available, skipping")

                    # Get the list of contents in the algorithm directory
                    algorithm_contents = os.listdir(content_path)
                    print("      Algorithm Contents:")
                    for outcome in algorithm_contents:
                        print("     ",outcome)
                        model_dic[algorithm][outcome] = {}
                        
                        print("           Outcome Contents:")
                        
                        '''
                        # Get the list of contents in the outcome directory
                        try:
                            outcome_contents = os.listdir(os.path.join(content_path, outcome))
                            print("           Outcome Contents:")
                            model_name = algorithm + '_' + outcome + '_model.joblib'
                            print("          ",model_name)
                            loaded_model = joblib.load(os.path.join(content_path, outcome, model_name))
                            print("          ",loaded_model)
                        except:
                            model_name = algorithm + '_' + outcome + '_model.joblib'
                            print("          ",model_name)
                            with open(os.path.join(content_path, outcome, model_name), "rb") as f:
                                loaded_model = pickle.load(f)
                            print("          ",loaded_model) 
                        '''
                        
                        # get everything else needed
                        outcome_contents = os.listdir(os.path.join(content_path, outcome))
                        print("           Outcome Contents:")
                        model_name = algorithm + '_' + outcome + '_model.joblib'
                        print("          ",model_name)
                        loaded_model = joblib.load(os.path.join(content_path, outcome, model_name))
                        model_dic[algorithm][outcome]['Model'] = loaded_model
                        print("          ",loaded_model)
                            
                        feature_names = algorithm + '_' + outcome + '_selected_features.joblib'
                        print("          ",feature_names)
                        loaded_features = joblib.load(os.path.join(content_path, outcome, feature_names))
                        model_dic[algorithm][outcome]['Features'] = loaded_features
                        #print("          ",loaded_features)
                            
                        numeric_cols = algorithm + '_' + outcome + '_numeric_cols.joblib'
                        print("          ",numeric_cols)
                        loaded_numeric_cols = joblib.load(os.path.join(content_path, outcome, numeric_cols))
                        model_dic[algorithm][outcome]['Numeric Columns'] = loaded_numeric_cols
                        #print("          ",loaded_numeric_cols)
                            
                        cat_cols = algorithm + '_' + outcome + '_cat_cols.joblib'
                        print("          ",cat_cols)
                        loaded_cat_cols = joblib.load(os.path.join(content_path, outcome, cat_cols))
                        model_dic[algorithm][outcome]['Categorical Columns'] = loaded_cat_cols
                        #print("          ",loaded_numeric_cols)

                        groundtruth_name = algorithm + '_' + outcome + '_true_fullset.joblib'
                        print("          ",groundtruth_name)
                        loaded_groundtruth = joblib.load(os.path.join(content_path, outcome, groundtruth_name))
                        model_dic[algorithm][outcome]['Outcome Name'] = loaded_groundtruth.name
                        print("           Outcome Name: ",loaded_groundtruth.name)

                        imputer_name = algorithm + '_' + outcome + '_imputer.joblib'
                        print("          ",imputer_name)
                        try:
                            loaded_imputer = joblib.load(os.path.join(content_path, outcome, imputer_name))
                            model_dic[algorithm][outcome]['Imputer'] = loaded_imputer
                            print("          ",loaded_imputer)
                        except:
                            print("          No Imputer")

                        scaler_name = algorithm + '_' + outcome + '_scaler.joblib'
                        print("          ",scaler_name)
                        try:
                            loaded_scaler = joblib.load(os.path.join(content_path, outcome, scaler_name))
                            model_dic[algorithm][outcome]['Scaler'] = loaded_scaler
                            print("          ",loaded_scaler)
                        except:
                            print("          No Scaler")

                        normalizer_name = algorithm + '_' + outcome + '_normalizer.joblib'
                        print("          ",normalizer_name)
                        try:
                            loaded_normalizer = joblib.load(os.path.join(content_path, outcome, normalizer_name))
                            model_dic[algorithm][outcome]['Normalizer'] = loaded_normalizer
                            print("          ",loaded_normalizer)
                        except:
                            print("          No Normalizer")

                        res_train_name = algorithm + '_' + outcome + '_res_table.joblib'
                        #print(res_train_name)
                        df_res_train = joblib.load(os.path.join(content_path, outcome, res_train_name))
                        model_dic[algorithm][outcome]['Train Set Res'] = df_res_train
                        print("           Train Set Res: ",df_res_train)

                        array_train_name = algorithm + '_' + outcome + '_res_array.joblib'
                        #print(array_train_name)
                        df_array_res_train = joblib.load(os.path.join(content_path, outcome, array_train_name))
                        model_dic[algorithm][outcome]['Train Set Res Array'] = df_array_res_train
                        #print("           Train Set Res Array: ",df_array_res_train)
                        st.write("We got Model Saved!")
 
                            
    return model_dic, input_cols_dic


def generate_results_table(results_dictonary):
    # Initialize an empty list to collect rows
    rows = []
    no_results = []
    
    for _, (algo, outcomes) in enumerate(results_dictonary.items()):

        for outcome in list(outcomes.keys()):
            print(f'{outcome} for {algo}')
            
            for _, (key, value) in enumerate(outcomes[outcome]['evaluation'].items()): 
                if isinstance(value, np.int32):
                    print("Int 32 Detected!")
                    print(f"It is on {key}")
        
            new_row = {'Outcome': outcome,
                      'Algorithm': algo,
                      'AUROC Score': outcomes[outcome]['evaluation']['AUROC Score'],
                      'AUROC CI Lower': outcomes[outcome]['evaluation']['AUROC CI Low'],
                      'AUROC CI Upper': outcomes[outcome]['evaluation']['AUROC CI High'],
                      'Accuracy': outcomes[outcome]['evaluation']['Accuracy'],
                      'Precision': outcomes[outcome]['evaluation']['precision'],
                      'Recall': outcomes[outcome]['evaluation']['recall'],
                      'TPR': outcomes[outcome]['evaluation']['TPR'], # same as Sensitivity 
                      'TNR': outcomes[outcome]['evaluation']['TNR'], # same as Specificity 
                      'FPR': outcomes[outcome]['evaluation']['FPR'], 
                      'FNR': outcomes[outcome]['evaluation']['FNR'],
                      'PPV': outcomes[outcome]['evaluation']['PPV'],
                      'NPV': outcomes[outcome]['evaluation']['NPV'],
                      'TP': outcomes[outcome]['evaluation']['TP'],
                      'FP': outcomes[outcome]['evaluation']['FP'],
                      'TN': outcomes[outcome]['evaluation']['TN'],
                      'FN': outcomes[outcome]['evaluation']['FN'],
                      'Cutoff value': outcomes[outcome]['evaluation']['cutoff'],
                      'P': outcomes[outcome]['evaluation']['P'],
                      'N': outcomes[outcome]['evaluation']['N'],
                      'AUROC Score (Train)': outcomes[outcome]['evaluation']['AUROC Score (Train)'],
                      'AUROC CI Lower (Train)': outcomes[outcome]['evaluation']['AUROC CI Low (Train)'],
                      'AUROC CI Upper (Train)': outcomes[outcome]['evaluation']['AUROC CI High (Train)'],
                      'P (Train)': outcomes[outcome]['evaluation']['P (Train)'],
                      'N (Train)': outcomes[outcome]['evaluation']['N (Train)']}

            # add new row
            rows.append(new_row)

    # Specifying the list of outcomes with no results for demonstration purposes
    print(f'Outcomes with no results are {no_results}')

    # Writing the output to a text file
    with open("outcomes_no_results.txt", "w") as file:
        file.write(f"Outcomes with no results are {no_results}")

    # Convert the list of rows into a DataFrame
    results_df = pd.DataFrame(rows)
    return results_df

def get_avg_metric(variable, metrics):
    values_list = []
    for metric in metrics:
        values_list.append(metric[variable])

    values_avg = sum(values_list) / len(values_list)
    return values_avg

# get the average metric scores in results_dictonary values
def get_avg_results_dic(results_dictonary):

    results_dictonary_avg = {}

    for _, (algo, data) in enumerate(results_dictonary.items()):
        #st.write(algo)
        #st.write(metrics)
        results_dictonary_avg[algo] = {}

        for _, (outcome, values) in enumerate(data.items()):
            #st.write(outcome)
            #t.write(values)
            metric_dic_train = {}
            metric_dic_test = {}

            results_dictonary_avg[algo][outcome] = {}
            metrics_train = values['Metrics']['Train']
            metrics_test = values['Metrics']['Test']

            metric_dic_test['TPR'] = get_avg_metric('TPR', metrics_test)
            metric_dic_test['TNR'] = get_avg_metric('TNR', metrics_test)
            metric_dic_test['FPR'] = get_avg_metric('FPR', metrics_test)
            metric_dic_test['FNR'] = get_avg_metric('FNR', metrics_test)

            metric_dic_test['PPV'] = get_avg_metric('PPV', metrics_test)
            metric_dic_test['NPV'] = get_avg_metric('NPV', metrics_test)

            metric_dic_test['AUROC Score'] = get_avg_metric('AUROC Score', metrics_test)
            metric_dic_test['AUROC CI Low'] = get_avg_metric('AUROC CI Low', metrics_test)
            metric_dic_test['AUROC CI High'] = get_avg_metric('AUROC CI High', metrics_test)

            metric_dic_test['cutoff'] = get_avg_metric('cutoff', metrics_test)
            metric_dic_test['cutoff type'] = metrics_test[0]['cutoff type']

            metric_dic_test['Accuracy'] = get_avg_metric('Accuracy', metrics_test)

            metric_dic_test['precision'] = get_avg_metric('precision', metrics_test)
            metric_dic_test['recall'] = get_avg_metric('recall', metrics_test)

            metric_dic_test['TP'] = get_avg_metric('TP', metrics_test)
            metric_dic_test['FP'] = get_avg_metric('FP', metrics_test)
            metric_dic_test['TN'] = get_avg_metric('TN', metrics_test)
            metric_dic_test['FN'] = get_avg_metric('FN', metrics_test)

            metric_dic_test['P'] = get_avg_metric('P', metrics_test)
            metric_dic_test['N'] = get_avg_metric('N', metrics_test)

            metric_dic_test['AUROC Score (Train)'] = get_avg_metric('AUROC Score', metrics_train)
            metric_dic_test['AUROC CI Low (Train)'] = get_avg_metric('AUROC CI Low', metrics_train)
            metric_dic_test['AUROC CI High (Train)'] = get_avg_metric('AUROC CI High', metrics_train)

            metric_dic_test['P (Train)'] = get_avg_metric('P', metrics_train)
            metric_dic_test['N (Train)'] = get_avg_metric('N', metrics_train)


            results_dictonary_avg[algo][outcome]['evaluation'] = metric_dic_test

            results_dictonary_avg[algo][outcome]['roc'] = {}
            results_dictonary_avg[algo][outcome]['roc']['Train'] = values['ROC']['Train']
            results_dictonary_avg[algo][outcome]['roc']['Test'] = values['ROC']['Test']

            results_dictonary_avg[algo][outcome]['conf_matrix'] = {}
            results_dictonary_avg[algo][outcome]['conf_matrix']['Train'] = values['Conf_Matrix']['Train'].tolist()
            results_dictonary_avg[algo][outcome]['conf_matrix']['Test'] = values['Conf_Matrix']['Test'].tolist()
    
    #st.write(results_dictonary_avg)
    return results_dictonary_avg

# create the configuration file
def generate_configuration_file(num_exp, project_name, train_set, test_sets, exp_names, algorithms, exp_type, options, param_vals):
    configuration_dic = {}
    configuration_dic[project_name] = {}
    #configuration_dic[project_name]['train_set']  = train_set
    #configuration_dic[project_name]['test_sets']  = test_sets
    configuration_dic[project_name]['exp_type']  = exp_type
    
    for i in range(num_exp):
        #st.write(i)
        experiment = {}
        experiment['algorithm'] = algorithms[i]
        #experiment['training_type'] = training_type[i]
        experiment['options'] = options[i]
        experiment['param_vals'] = param_vals[i]
        
        configuration_dic[project_name][exp_names[i]] = experiment
    
    #st.write("Configuration_dic : ", configuration_dic)
    
    return configuration_dic
    

# generate configuration template
def generate_configuration_template(project_name, num_exp, training_method):
    #st.write("Number of exp. :", num_exp)
    exp_names=["exp_" + str(i) for i in range(num_exp)]
    algorithms=["enter_algo_here"]*num_exp
    train_set="enter_filename_here"
    options=[options_test_set]*num_exp if training_method=="Dedicated Test Set" else [options_default]*num_exp
    
    configuration_dic = generate_configuration_file(num_exp=num_exp, 
                            project_name=project_name, 
                            train_set=train_set,
                            test_sets=["None"],
                            exp_names=exp_names,
                            algorithms=algorithms, 
                            exp_type="enter_type_here", 
                            #training_type=["enter_type_here"] * num_exp,
                            options=options,
                            param_vals=["None"]*num_exp)

    #st.write(configuration_dic)

    return configuration_dic

def generate_congfig_file(exp_name, algorithims, exp_type, options):
    print("Number of exp. :", len(algorithims))

    configuration_dic = {}
    configuration_dic[exp_name] = {}
    configuration_dic[exp_name]['exp_type']  = exp_type

    for algorithim in algorithims:
        experiment = {}
        experiment['algorithm'] = algo_shortnames[algorithim]
        experiment['options'] = options
        experiment['param_vals'] = "None"

        configuration_dic[exp_name][f"exp_{algorithim}"] = experiment
    
    #st.write("Configuration_dic : ", configuration_dic)
    
    return configuration_dic