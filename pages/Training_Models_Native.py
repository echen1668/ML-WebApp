from pathlib import Path
import time
import io
import pandas as pd
import numpy as np
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn import metrics
import time
import psutil
import autogluon
import pandas as pd
import numpy as np
import json
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
#from scipy import interp
from scipy.stats import norm
from scipy.special import ndtri
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import magic 
import pickle
import random 
from random import randint
from random import uniform
from scipy import stats
import ast
from pathlib import Path
import joblib as joblib
from joblib import dump, load
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from sklearn.model_selection import cross_validate
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import RocCurveDisplay
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.metrics import auc
from sklearn.utils import resample
from sklearn.model_selection import KFold
from sklearn.model_selection import StratifiedKFold, RepeatedStratifiedKFold
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import VarianceThreshold
from sklearn.model_selection import RandomizedSearchCV, GridSearchCV
from sklearn.calibration import CalibratedClassifierCV
from sklearn.calibration import calibration_curve, CalibrationDisplay
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score
from sklearn.metrics import classification_report
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import label_binarize
from sklearn.metrics import confusion_matrix
#from scipy import interp
from scipy.stats import norm
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
import random 
from random import randint
from random import uniform
from scipy import stats
import mrmr
from mrmr import mrmr_classif
import xgboost
import catboost
import shap
from scipy import stats
import os
import joblib as joblib
from joblib import dump, load
import json
import tkinter as tk
from tkinter import *
from autogluon.tabular import TabularDataset, TabularPredictor
import mimetypes
from difflib import SequenceMatcher
#np.random.seed(1000)
rstate = 12
from datetime import datetime
# import module
import pprint
import pymongo
from pymongo import MongoClient
import streamlit as st
from streamlit_cookies_manager import EncryptedCookieManager
from Multi_Outcome_Classification_tools import multi_outcome_hyperparameter_binary, multi_outcome_hyperparameter_binary_train_and_test, multi_outcome_cv
from Common_Tools import generate_configuration_file, generate_configuration_template, generate_results_table, generate_congfig_file, get_avg_results_dic, wrap_text_excel, expand_cell_excel, grid_excel, generate_all_idx_files, upload_data, load_data, save_data, data_prep, data_prep_train_set, parse_exp_multi_outcomes, setup_multioutcome_binary, refine_binary_outcomes, generate_joblib_model
from roctools import full_roc_curve, plot_roc_curve

algo_shortnames = { # short names for ML Algorithims
    "Random Forest": "rf",
    "XGBoost": "xgb",
    "Cat Boost": "cat",
    "SGD Elastic": "sgd_elastic",
    "SGD L2": "sgd_l2",
    "Logistic Reg. L2": "lr_l2",
    "Logistic Reg.": "lr", 
    "Descision Tree": "dt", 
    "SVM": "svm",
    "KNearest N.": "knn", 
}

option_names = { # potenital values in put in the options section of the configuartion file
    "impute_strategy": "['mean', 'median', 'most_frequent', 'constant']",
    'cut threshold (range)': "(0.0 - 1.0)",
    "inf (how to handle inf values in data)": "['replace with null', 'replace with zero']",
    "outliers (how to handle outliers in data)": "['None', 'remove rows', 'log']",
    "scalingMethod": "['MinMaxScaler', 'RobustScaler', 'MaxAbsScaler', 'StandardScaler']",
    "rebalance_type": "['RandomUnderSampler', 'RandomOverSampler', 'SMOTE', 'ADASYN']",
    "sampling_strategy": "['auto', 'majority', 'not minority', 'not majority', 'all', 'ratio']",
    "FeatureSelection_method": "['MRMR', 'RFECV', 'SelectKBest-f_classif', 'SelectKBest-chi2']", 
    "strategy": "['random', 'bayesian', 'grid']"
}

# --- Page Configuration ---
st.set_page_config(
    page_title="(Sklearn) Create a New Experiment",
    page_icon="🔥",
    layout="wide"
)

data_sets = {}

# back button to return to main page
if st.button('Back'):
    st.switch_page("pages/Training_Models_Options.py")  # Redirect to the main back 

# --- UI Initialization ---
st.title("🚀-🔥 Create a New Experiment (Sklearn)")
st.markdown("Configure and launch a new model training workflow using the traditional/native Sklearn framework.")

# Check if client_name was passed
cookies = EncryptedCookieManager(prefix="mlhub_", password="some_secret_key")
if not cookies.ready():
    st.stop()

# Check cookies first
if "client_name" in cookies:
    st.session_state["client_name"] = cookies["client_name"]
    #st.write(cookies["client_name"])
#else:
    #st.error("No found")

# then check in session state
if "client_name" not in st.session_state:
    st.error("No database connection found. Please go back to the main page.")
    st.stop()

client_name = st.session_state["client_name"]


# connect to database
#client = MongoClient('10.14.1.12', 27017)
client = MongoClient(client_name, 27017)

# create the database if it does not already exists
db = client.machine_learning_database

# create tables for models in the databse
models = db.models

# create the results if it does not already exists
results = db.results

# create the results if it does not already exists
datasets = db.datasets

# get all unique exp. names from results collection
exp_names_n = db.models.distinct("exp_name")
exp_names_cv = db.results.distinct("exp_name", {"type": "Native-CV"})
exp_names = exp_names_n + exp_names_cv

# get all training data names from database
data_names_train = db.datasets.distinct("data_name", {"type": "Train"})

# get all testing data names from database
data_names_list_test = db.datasets.distinct("data_name", {"type": "Test"})

# run training
def project(configuration_dic, data_sets, unique_value_threshold=10):

    st.write("---")
    st.subheader("🚀 Starting Experiment...")

    # --- 1. Setup Folders ---
    with st.spinner("Setting up project structure..."):
        # get name of the project as part of file name
        project_name = list(configuration_dic.keys())[0]
        project_folder = os.path.join("Models", project_name)
        os.makedirs(project_folder, exist_ok=True)  # Create folder for algorithm results

        train_set = data_sets["Training Set"]
        index_set = data_sets["Index Set"]
        #test_sets = data_sets["Testing Set"]

        # save the data sets
        #save_data(train_set['Name'], train_set['Data'], os.path.join(project_folder, train_set['Name']))
        save_data(index_set['Name'], index_set['Data'], os.path.join(project_folder, index_set['Name']))

        if train_set['Name'] not in data_names_train:
            st.info(f"Training Dataset {train_set['Name']} is saving in the database", icon="ℹ️")
            # create a list of ML exp.'s that the dataset was used on
            exp_list = [project_name]
            # get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
            # save train set into data folder and database
            os.makedirs("Data Sets", exist_ok=True)
            save_data(train_set['Name'], train_set['Data'], os.path.join("Data Sets", train_set['Name']))
            dataset_train = {
                    "data_name": train_set['Name'],
                    "type": "Train",
                    "time_saved": current_time,
                    "data_path": os.path.join("Data Sets", train_set['Name']),
                    "exps used": exp_list
            }
            datasets.insert_one(dataset_train)
        else:
            st.info(f"Training Dataset {train_set['Name']} of the same name is already in the database. Will be overwritten in the database", icon="ℹ️")
            save_data(train_set['Name'], train_set['Data'], os.path.join("Data Sets", train_set['Name']))
            # update the dataset in datbase to trackdown the list of ML exps the set was used on
            dataset = datasets.find_one({"data_name": train_set['Name'], "type": "Train"})
            # Get the current list of experiments or initialize it if not present
            exp_list = dataset.get("exps used", [])
            # Add the current project name if it's not already in the list
            if project_name not in exp_list:
                exp_list.append(project_name)

            datasets.update_one(
                {"data_name": train_set['Name'], "type": "Train"}, # Filter condition
                {"$set": { "exps used": exp_list }} # Update operation
            )

        # get proper data name
        if train_set['Name'].endswith('.xlsx'):
            data_name = train_set["Name"][:-len('.xlsx')]
        elif train_set["Name"].endswith('.csv'):
            data_name = train_set["Name"][:-len('.csv')]
        else:
            data_name = train_set["Name"]

        # get input and ouput columns
        input_cols, label_cols, categorical_cols, numeric_cols = parse_exp_multi_outcomes(train_set["Data"], index_set["Data"], unique_value_threshold=unique_value_threshold)
        st.write("Label Columns: ", label_cols)
        #st.write("Categorical Columns: ", categorical_cols)
        #st.write("Numeric Columns: ", numeric_cols)
        input_cols_og = input_cols.copy() # keep the origional input list

        threshold_type = configuration_dic[project_name]['threshold_type']
        df_train = refine_binary_outcomes(train_set["Data"], label_cols)

    algorithms = [] # list of algorithims
    results_dictonary = {}
    for experiment in configuration_dic[project_name]:
        experiment_name = experiment
        if experiment_name in ["train_set", "test_sets", "threshold_type", "exp_type"]:
            continue

        # --- 2. Prep the data ---
        st.write(f"Experiment Name: {experiment_name} had started training...")

        with st.spinner(f"Preping the data for {experiment_name}..."):
            set_up = configuration_dic[project_name][experiment_name]
            options, algorithm, param_vals = setup_multioutcome_binary(set_up, experiment_name, project_folder)

            algorithms.append(algorithm)
            results_dictonary[algorithm] = {}

            # refine the training set before model training
            df_train_refined, input_cols, label_cols, encoder, encoded_cols, qt = data_prep(df_train, input_cols, label_cols, numeric_cols, categorical_cols, options)
            #st.write(df_train_refined)
            # save any preprocessing steps
            algorithm_folder = os.path.join(project_folder, experiment_name)
            os.makedirs(algorithm_folder, exist_ok=True)  # Create folder for algorithm preprocessing steps

            if encoder != None:
                encoder_name = os.path.join(algorithm_folder, algorithm + "_encoder.joblib")
                joblib.dump(encoder, encoder_name)
                encoded_cols_name = os.path.join(algorithm_folder, algorithm + "_encoded_cols.joblib")
                joblib.dump(encoded_cols, encoded_cols_name)

            if qt != None:
                qt_name = os.path.join(algorithm_folder, algorithm + "_qt.joblib")
                joblib.dump(qt, qt_name)

        # --- 3. Loop Through Algorithms and Execute Workflows ---
        with st.spinner(f"Training has offically started for {experiment_name}! ... This may take a while."):

            if st.session_state.training_method == "Train Whole Set":
                multi_outcome_hyperparameter_binary(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, options, algorithm, param_vals, experiment_name, project_folder)
                
            elif st.session_state.training_method == "Train/Test Split":
                algo_dictonary = multi_outcome_hyperparameter_binary_train_and_test(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, threshold_type, options, algorithm, param_vals, experiment_name, project_folder, project_name, data_name)
                results_dictonary[algorithm] = algo_dictonary
            
            elif st.session_state.training_method == "Cross-Validation":
                algo_dictonary = multi_outcome_cv(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, threshold_type, options, algorithm, param_vals, experiment_name, project_folder, project_name)
                results_dictonary[algorithm] = algo_dictonary

        st.success(f"Successfully completed training for **{experiment_name}**.")
    
    
    st.success(f"**Successfully completed training for all experiments!**")

    # --- 4. Finalize Experiment ---
    st.write("---")
    with st.spinner("Finalizing experiment: saving metadata and final reports..."):

        if st.session_state.training_method in ["Train/Test Split", "Train Whole Set"]:
            models_dic, _ = generate_joblib_model(project_folder)
            model_absolute_path = os.path.join(project_folder)

            pathway_name = os.path.join(model_absolute_path, project_name + "_models.joblib")
            joblib.dump(models_dic, pathway_name)

            # get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
            
            model = {
                "exp_name": project_name,
                "type": "Native",
                "model_path": pathway_name,
                "algorithms": algorithms,
                "input variables": input_cols,
                "input variables (original)": input_cols_og,
                'outcomes': label_cols,
                "configuration": configuration_dic,
                "train_data": train_set["Name"],
                "time_created": current_time
            }

            models.insert_one(model) # insert one dictonary 
        
        if st.session_state.training_method in ["Train/Test Split", "Cross-Validation"]:

            # save the results into database
            results_folder = os.path.join("Results", project_name, data_name) if st.session_state.training_method=="Train/Test Split" else os.path.join("Results", project_name)
            os.makedirs(results_folder, exist_ok=True)  # Create folder for algorithm results if it doesn't exists yet
            filename = os.path.join(results_folder, "metadata.txt")
            f = open(filename, "w", encoding="utf-8")
            f.write("\nExp Name: %s"% project_name)
            f.write("\nInput Columns: %s"% input_cols_og)
            f.write("\nOutput Columns: %s"% label_cols)
            f.write("\nAlgorithms: %s"% algorithms)
            f.close()
            

            training_type = "Native" if st.session_state.training_method=="Train/Test Split" else "Native-CV"

            final_results_dic = get_avg_results_dic(results_dictonary, override=True) if st.session_state.training_method=="Cross-Validation" else results_dictonary

            path_name = os.path.join(results_folder, f"{project_name}_results.joblib")
            joblib.dump(final_results_dic, path_name)

            # generate the results table
            results_df = generate_results_table(final_results_dic)
            table_name = os.path.join(results_folder, f"{project_name}_results.xlsx")

            results_df.to_excel(table_name, index=False, engine='openpyxl')

            results_df.to_excel(table_name, index=False)
            results_df.to_excel(table_name, index=False)
            expand_cell_excel(table_name)
            wrap_text_excel(table_name)
            grid_excel(table_name)

            # Convert results_df to list of dictionaries
            results_dic = results_df.to_dict(orient='records')

            # get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

            try:
                result = {
                    "exp_name": project_name,
                    "type": training_type,
                    "test set": data_name,
                    "threshold used": threshold_type,
                    "results_dic": final_results_dic,
                    "results_table": results_dic,
                    'dataset used': train_set['Name'],
                    "algorithms": algorithms,
                    "input variables": input_cols,
                    "input variables (original)": input_cols_og,
                    'outcomes': label_cols,
                    "time_created": current_time
                }

                if st.session_state.training_method == "Cross-Validation": # add configuration if the experiment is cross validation
                    result['configuration'] = configuration_dic

                results.insert_one(result) # insert one dictonary
            except:
                st.info("Results size is too large. Will save filepaths instead.")

                result = {
                    "exp_name": project_name,
                    "type": training_type,
                    "test set": data_name,
                    "threshold used": threshold_type,
                    "results_dic": path_name,
                    "results_table": results_dic,
                    'dataset used': train_set['Name'],
                    "algorithms": algorithms,
                    "input variables": input_cols,
                    "input variables (original)": input_cols_og,
                    'outcomes': label_cols,
                    "time_created": current_time
                }

                if st.session_state.training_method == "Cross-Validation": # add configuration if the experiment is cross validation
                    result['configuration'] = configuration_dic

                results.insert_one(result) # insert one dictonary
    
    st.success(f"✅ Experiment '{project_name}' completed successfully!")
    
    if st.session_state.training_method in ["Cross-Validation"]:
        st.subheader("Jump to Visualizing Results") # redirect to the testing section
        st.page_link("pages/Visualize_Multi_Results (CV).py", label="Visualize Results", icon="📊")
    else:
        st.subheader("Jump to Visualizing Results") # redirect to the testing section
        st.page_link("pages/Visualize_Multi_Results (Native).py", label="Visualize Results", icon="📊")
        st.write("or")
        st.subheader("Jump to Testing the Models") # redirect to the testing section
        st.page_link("pages/Testing_Native_Models.py", label="Test Models", icon="🧪")


def create_column_summary(df, index_df): # get a summry of the dataframe
    # get the selected columns
    inputs = index_df.columns[index_df.iloc[0] == 1].tolist()
    #st.write(inputs) 

    summary_list = []
    for col in df[inputs].columns:
        col_type = df[col].dtype
        missing_count = df[col].isnull().sum()
        missing_percent = f"{missing_count / len(df) * 100:.2f}%"
        unique_count = df[col].nunique()
        
        col_summary = {
            "Column": col, "Data Type": str(col_type), "Missing Values": missing_count,
            "Missing (%)": missing_percent, "Unique Values": unique_count
        }
        
        if np.issubdtype(col_type, np.number):
            stats = df[col].describe()
            col_summary.update({
                "Mean": f"{stats.get('mean', 0):.2f}", "Std Dev": f"{stats.get('std', 0):.2f}",
                "Min": stats.get('min', 0), "Max": stats.get('max', 0)
            })
        summary_list.append(col_summary)
    return pd.DataFrame(summary_list)

def create_experiment_summary(data_df, index_df): #create a summery of the experiment
    summary_list = []
    for exp_name, row in index_df.iterrows():
        inputs = [col for col, val in row.items() if val == 1]
        outcomes = [col for col, val in row.items() if val == 2]
        
        for outcome in outcomes:
            prevalence = "N/A"
            if outcome in data_df.columns and set(data_df[outcome].dropna().unique()).issubset({0, 1}):
                prev_rate = data_df[outcome].value_counts(normalize=True).get(1, 0)
                prevalence = f"{prev_rate * 100:.2f}%"
            summary_list.append({
                "Experiment Plan": exp_name, "Input Count": len(inputs),
                "Outcome": outcome, "Prevalence": prevalence
            })
    return pd.DataFrame(summary_list)

# gnerate an index file
def gen_idx(df_train, train_set_name, nexp:int=1, sheets:list=None):

    if train_set_name.endswith('.xlsx'):
        filename = train_set_name[:-len('.xlsx')]
    elif train_set_name.endswith('.csv'):
        filename = train_set_name[:-len('.csv')]
    else:
        filename = train_set_name

    indexFileName = filename + '_index.xlsx'
    
    all_columns = list(df_train.columns)

    df_index = pd.DataFrame(columns=all_columns)
    df_index.loc['Key'] = 0

    # convert df_index to an Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_index.to_excel(writer, index=True)

    st.download_button(
        label="Download the Index File Template to set inputs and outputs ⬇️",
        data=output.getvalue(),
        file_name=indexFileName,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return indexFileName

# (The rest of the UI code is the same as before)
# ...
if 'training_method' not in st.session_state:
    st.session_state.training_method = "Train/Test Split"
    st.session_state.exp_name = ""
    st.session_state.algorithms = []

# --- Step 1: Experiment Setup ---
st.header("Step 1: Define Experiment and Data Strategy")
project_name = st.text_input("Experiment Name", help="Enter a unique name for this experiment.")

#st.write(exp_names)
if project_name in exp_names:
    st.error("Experiment with that name already exists")
    is_valid = True
    configuration_dic = None
elif project_name is None or project_name=="" or project_name.isspace():
    st.info("Please enter a name for the experiment")
    is_valid = True
    configuration_dic = None
elif '/' in project_name or '\\' in project_name:
    st.error("Invalid Experiment Name (Cannot have / or \\)")
    is_valid = True
    configuration_dic = None
else:
    is_valid = False

st.radio(
    "Select Training Method",
    ["Train/Test Split", "Train Whole Set", "Cross-Validation"],
    key="training_method",
    horizontal=True,
    help="Choose how to evaluate model."
)

# --- Step 2: Upload Data (Dynamic UI) ---
st.header("Step 2: Upload Data")

#user chooses whatever to upload the data or retrive a past data set from the database
data_options = st.radio("Choose an option:", ["Upload a dataset", "Retrive dataset from database"])

if data_options == "Upload a dataset":
    data_name_train = None
    #data_names_test = []
    col1, col2 = st.columns(2)
    with col1:
        # Use a consistent key for the main dataset uploader
        dataset_uploader_key = "main_dataset_uploader"
        
        if st.session_state.training_method in ["Train/Test Split", "Cross-Validation"]:
            main_uploaded_file = st.file_uploader("Upload Dataset (CSV or Excel)", type=['csv', 'xlsx'])
            #test_uploaded_file = None
        else: # Train Whole Set
            main_uploaded_file = st.file_uploader("Upload Training Dataset", type=['csv', 'xlsx'], key=dataset_uploader_key)
            #test_uploaded_file = st.file_uploader("Upload Testing Dataset", type=['csv', 'xlsx'], key="test_dataset", accept_multiple_files=True)

    with col2:
        completed_index_file = st.file_uploader("Upload **Completed** Index File", type=['xlsx'])
else:
    main_uploaded_file = None
    #test_uploaded_file = None

    # Dropdown to select the training dataset
    data_name_train = st.selectbox("Select a Training Dataset from the database:", data_names_train, index=None, placeholder="Select One...")

    #if st.session_state.training_method == "Train Whole Set":
        # Dropdown to select the tesitng dataset
    #    data_names_test = st.multiselect("Select a Testing Dataset from the database:", data_names_list_test)
    #else:
    #    data_names_test = []

    # Open and wrap the file like an uploaded file (binary mode)
    if data_name_train is not None:
        completed_index_file = st.file_uploader("Upload **Completed** Index File", type=['xlsx'])
       


st.write("")

#st.write(main_uploaded_file)
#st.write(test_uploaded_file)


# upload the dataset(s)
if data_options == "Upload a dataset" and main_uploaded_file:
    train_set = main_uploaded_file.name
    #st.write(train_set)

    df_train = load_data(train_set, main_uploaded_file)

    st.write(df_train)  # Display the first few rows

    data_sets["Training Set"] = {}
    data_sets["Training Set"]["Name"] = main_uploaded_file.name
    data_sets["Training Set"]["Data"] = df_train
elif data_name_train:
    df_train = upload_data(os.path.join("Data Sets",data_name_train))
    st.write(df_train)  # Display the first few rows
    data_sets["Training Set"] = {}
    data_sets["Training Set"]["Name"] = data_name_train
    data_sets["Training Set"]["Data"] = df_train


#test_sets = []
#if data_options == "Upload a dataset" and test_uploaded_file:

#   data_sets["Testing Set"] = {}

#    for file in test_uploaded_file:
        #st.subheader(f"Dataset: {file.name}")
#        test_sets.append(file.name)

#        df_test = load_data(file.name, file)
        #st.write(df_test.head())  # Display the first few rows

#        data_sets["Testing Set"][file.name] = df_test
#elif len(data_names_test) > 0:
#    data_sets["Testing Set"] = {}

#    for data_name_test in data_names_test:
#        test_sets.append(data_name_test)
        #st.write(f"Dataset: {file.name}")
#        df_test = upload_data(os.path.join("Data Sets",data_name_test))
        #st.write(df_test.head())  # Display the first few rows
#        data_sets["Testing Set"][data_name_test] = df_test

#else:
#    data_sets["Testing Set"] = {}

#st.write(test_sets)

if (main_uploaded_file or data_name_train) and not completed_index_file:
    st.info("A dataset has been uploaded. Now generate a matching index file to edit.")
    gen_idx(df_train, data_sets["Training Set"]["Name"])

    
# Add feedback to the user once they've uploaded their completed file.
elif (main_uploaded_file or data_name_train) and completed_index_file:

    # upload the index file
    index_set = pd.read_excel(completed_index_file)
    #st.write(index_set.head())  # Display the first few rows

    data_sets["Index Set"] = {}
    data_sets["Index Set"]["Name"] = completed_index_file.name
    data_sets["Index Set"]["Data"] = index_set

    #st.success("✅ Completed index file has been uploaded.")
    if st.button("📊 View Data Summary and Experiment Plan", use_container_width=True):
        # When button is clicked, open a dialog
        st.dialog("Data Summary")
        st.header("Experiment Plan Summary")
        st.markdown("Table: summarizes the inputs and outcomes.")
            
        exp_summary_df = create_experiment_summary(df_train, index_set)
        st.dataframe(exp_summary_df, use_container_width=True)

        st.header("Dataset Column Analysis")
        st.markdown("Table: Statistics for each column in the dataset.")
        col_summary_df = create_column_summary(df_train, index_set)
        st.dataframe(col_summary_df, use_container_width=True)

        if st.button("Close"):
            st.rerun() # Closes the dialog

# --- Step 3: Configure Training Pipeline ---
st.header("Step 3: Configure Training Pipeline")

if (main_uploaded_file or data_name_train) and completed_index_file and is_valid==False:
    configuration_dic = None
    # choose how you configure your model(s)
    configure_options = st.radio("Choose an option:", ["Upload a file", "User Customization"])
else:
    configure_options = None
    configuration_dic = None

if configure_options == "User Customization": #st.session_state.get("main_dataset") or st.session_state.get("train_dataset"):
    
    algorithms = st.multiselect(
        "Select ML Algorithms to Train",
        ['Random Forest', 'XGBoost', 'Cat Boost', "SGD Elastic", "SGD L2", "Logistic Reg.", "Logistic Reg. L2", "Descision Tree", "SVM", "KNearest N."]
    )

    unique_value_threshold = st.number_input("Enter the minimum unique value threshold for a numerical input variable to be consired catagorical:", min_value=1, max_value=100, value=10)

    with st.expander("▶️ Data Preprocessing Options"):
        oneHotEncode = st.radio("One-Hot Encode Categorical Features?", ("True", "False"), horizontal=True)
        impute = st.radio("Impute Missing Values?", ("True", "False"), horizontal=True)
        if impute=="True":
            impute_strategy = st.selectbox("Imputing Strategy", ['mean', 'median', 'most_frequent', 'constant'])
        else:
            impute_strategy = "None"

        inf_handling = st.radio("Handle Infinite Values By:", ('replace with null', 'replace with zero'), horizontal=True)
        
        st.subheader("Handle Outliers")
        outliers = st.radio("Outlier Strategy:", ('None', 'remove rows', 'log'))
        if outliers != 'None':
            outliers_N = st.number_input("Outlier Threshold", value=50000)
        else:
            outliers_N = 0

        st.subheader("Handle Missing Rows")
        cutMissingRows = st.radio("Cut Rows with Missing Values?", ("True", "False"), horizontal=True)
        if cutMissingRows == 'True':
            cut_threshold = st.slider("Minimum % of Non-Null Values Required", 0.0, 1.0, 0.6)
        else:
            cut_threshold = None
            
    with st.expander("▶️ Feature Scaling & Transformation"):
        scaling = st.radio("Scale Numerical Features?", ("True", "False"), index=1, horizontal=True)
        if scaling == 'True':
            scalingMethod = st.selectbox("Scaling Method", ['MinMaxScaler', 'RobustScaler', 'MaxAbsScaler', 'StandardScaler'])
        else:
            scalingMethod = "None"
        
        QuantileTransformer = st.radio("Apply Quantile Transformation?", ("True", "False"), index=1, horizontal=True)
        normalize = st.radio("Apply Normalization?", ("True", "False"), index=1, horizontal=True)

    with st.expander("▶️ Rebalancing & Feature Selection"):
        rebalance = st.radio("Rebalance the Training Data?", ("True", "False"), index=1, horizontal=True)
        if rebalance == 'True':
            rebalance_type = st.selectbox("Rebalancing Method", ['RandomUnderSampler', 'RandomOverSampler', 'SMOTE', 'ADASYN'])
            sampling_strategy = st.selectbox("Sampling Strategy", ['auto', 'majority', 'not minority', 'not majority', 'all', 'ratio'])

            if sampling_strategy == "ratio":
                sampling_ratio = st.number_input("Ratio of # of samples in the minority class over # of samples in the majority class (should be calcauted as: n minority sample/N majority sample. Ex: have majority to outnumer minority 5 to 1 should calucate 0.20):", min_value=0.0, value=0.2, help="(should be calcauted as: n minority sample/N majority sample. Ex: have majority to outnumer minority 5 to 1 should calucate 1/5)")
            else:
                sampling_ratio = 0.0

            if rebalance_type in ['ADASYN', 'SMOTE']:
                k_neighbors = st.number_input("Number of Nearest Neighbors", min_value=1, value=5, help="The nearest neighbors used to define the neighborhood of samples to use to generate the synthetic samples.")
            else:
                k_neighbors = 5
        else:
            rebalance_type = "None"
            sampling_strategy = "None"
            sampling_ratio = 0.0
            k_neighbors = 5


        featureSelection = st.radio("Perform Feature Selection?", ("True", "False"), index=1, horizontal=True)
        if featureSelection == 'True':
            featureSelection_method = st.selectbox("Feature Selection Method", ['MRMR', 'RFECV', 'SelectKBest-f_classif', 'SelectKBest-chi2'])
            N_features = st.number_input("Number of Features to Select", min_value=1, value=20)
        else:
            featureSelection_method = "None"
            N_features = 0

    with st.expander("▶️ Training & Evaluation Options"):
        search_strategy = st.selectbox("Hyperparameter Search Strategy", ['random', 'bayesian', 'grid'])
        num_itr = st.number_input("Number of Search Iterations", min_value=1, value=50)
        k_fold = st.number_input("Number of Cross-Validation Folds", min_value=2, value=5)
        n_repeats = st.number_input("Number of CV Repeats", min_value=1, value=1)
        min_postives = st.number_input("Minimum Positive Cases for an Outcome", min_value=1, value=10)

        if st.session_state.training_method != "Train Whole Set":
            test_size = st.number_input("(for Train/Test Split Only) Enter the size of the test set (% of the data set set aside for testing):", min_value=0.0, max_value=1.0, value=0.2)
            threshold_type = st.selectbox("Optimal Cutoff Threshold Method", ['youden', 'mcc', 'ji', 'f1'])
        else:
            test_size = 0
            threshold_type = "None"

    options = {
        'oneHotEncode' : oneHotEncode, 
        'Impute': impute, 
        'impute_strategy': impute_strategy,
        'cutMissingRows' : cutMissingRows,
        "cut threshold": cut_threshold,
        "inf": inf_handling,
        'outliers': outliers, 
        'outliers_N': outliers_N, 
        'Scaling': scaling, 
        'scalingMethod': scalingMethod, 
        'QuantileTransformer': QuantileTransformer, 
        'Normalize': normalize,
        'rebalance' : rebalance,
        'rebalance_type': rebalance_type,
        'sampling_strategy': sampling_strategy,
        'sampling_ratio': sampling_ratio,
        'k_neighbors': k_neighbors,
        'FeatureSelection': featureSelection,
        "method": featureSelection_method,
        'N_features': N_features, 
        'strategy': search_strategy, 
        'itr': num_itr,
        "CV": k_fold,
        "n_repeats": n_repeats,
        "min_postives": min_postives,
        "test_size": test_size,
    }

    configuration_dic = generate_congfig_file(project_name, algorithms, threshold_type, options)    

elif configure_options == "Upload a file":
    unique_value_threshold = st.number_input("Enter the minimum unique value threshold for a numerical input variable to be consired catagorical:", min_value=1, max_value=100, value=10)
    # Number input
    num_models = st.number_input("Enter the number of models:", min_value=1, max_value=100, value=5)
    # reset configuration_dic
    configuration_dic = None
    # sends user to a seperate page where they can download an empty configuration file where they can customize themselves
    st.download_button(
        label="Download Configuration File Template ⬇️",
        data=json.dumps(generate_configuration_template(project_name, num_models), indent=4),
        file_name=f"{project_name}.json",
        mime="application/json"
    )

    st.subheader("Guide for setting up configuration file")

    with st.expander("▶️ List of ML Algorithims to use."): # provide the user a guide to all the ML alogrhtims that they can use.
        st.write("Please use the short versions of the algo names shown on the **right**.")
        st.write(algo_shortnames)


    with st.expander("▶️ Setup Options."): # provide the user a list of possible values to insert into the respective key in the configuration file
        st.write("Please use the names shown on the **lists**.")
        st.json(option_names)

    # File uploader for the training
    uploaded_config_file= st.file_uploader("Upload a Configuration File")

    if uploaded_config_file:
        configuration_dic = json.load(uploaded_config_file)
        config_name = list(configuration_dic.keys())[0] # get the name of model configuration and check if the model already exists in the database

        if config_name != project_name:
            st.write("Configuration File Name does not match Experiment Name!")
            is_exp_match = False
        else:
            is_exp_match = True
else:
    configuration_dic = None

# --- Step 4: Execute ---
st.header("Step 4: Run Experiment")

if configuration_dic != None and st.button("Start Training", type="primary", use_container_width=True):
    # Validation
    options = {key: val for key, val in st.session_state.items()}
        
    # Call the main orchestration function
    project(configuration_dic, data_sets, unique_value_threshold=unique_value_threshold)
