from pathlib import Path
import time
import io
import pandas as pd
import numpy as np
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn import metrics
import time
import psutil
import autogluon
import pandas as pd
import numpy as np
import json
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
#from scipy import interp
from scipy.stats import norm
from scipy.special import ndtri
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import magic 
import pickle
import random 
from random import randint
from random import uniform
from scipy import stats
import ast
from pathlib import Path
import joblib as joblib
from joblib import dump, load
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from sklearn.model_selection import cross_validate
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import RocCurveDisplay
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.metrics import auc
from sklearn.utils import resample
from sklearn.model_selection import KFold
from sklearn.model_selection import StratifiedKFold, RepeatedStratifiedKFold
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import VarianceThreshold
from sklearn.model_selection import RandomizedSearchCV, GridSearchCV
#from tune_sklearn import TuneSearchCV
from skopt import BayesSearchCV
from sklearn.calibration import CalibratedClassifierCV
from sklearn.calibration import calibration_curve, CalibrationDisplay
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score
from sklearn.metrics import classification_report
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import label_binarize
from sklearn.metrics import confusion_matrix
#from scipy import interp
from scipy.stats import norm
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
import random 
from random import randint
from random import uniform
from scipy import stats
import mrmr
from mrmr import mrmr_classif
import xgboost
import catboost
import shap
from scipy import stats
import os
import sys
import joblib as joblib
from joblib import dump, load
import json
import tkinter as tk
from tkinter import *
from autogluon.tabular import TabularDataset, TabularPredictor
from autogluon.common import space
from autogluon.common.space import Int, Real, Categorical
from difflib import SequenceMatcher
#np.random.seed(1000)
rstate = 12


# import module
from datetime import datetime
import pprint
import pymongo
from pymongo import MongoClient
import streamlit as st
from streamlit_cookies_manager import EncryptedCookieManager
from Multi_Outcome_Classification_tools import multi_outcome_hyperparameter_binary, multi_outcome_hyperparameter_binary_train_and_test, multi_outcome_cv
from Common_Tools import sanitize_filename, generate_configuration_file, generate_configuration_template, generate_results_table, generate_congfig_file, get_avg_results_dic, wrap_text_excel, expand_cell_excel, grid_excel, generate_all_idx_files, upload_data, load_data, save_data, data_prep, data_prep_train_set, parse_exp_multi_outcomes, setup_multioutcome_binary, refine_binary_outcomes, generate_joblib_model
from roctools import full_roc_curve, plot_roc_curve

algo_shortnames = { # short names for ML Algorithims
    "Random Forest": "RF",
    "XGBoost": "XGB",
    "Cat Boost": "CAT",
    "Extremely randomized trees": "XT", 
    "LightGBM": "GBM",
    "KNearest N.": "KNN", 
}

# custom models with custom hyperparameters
global custom_hyperparameters_sample
custom_hyperparameters_sample = {
    'GBM': {  # LightGBM
        'num_boost_round': space.Int(50, 400), 
        'learning_rate': space.Real(0.01, 0.1),  # Search space for learning rate
    },
    'XGB': {  # XGBoost
        'n_estimators': space.Int(10, 100),
        'max_depth': space.Int(2, 6),
        'colsample_bytree': space.Real(0.5, 0.8),
        'eta': space.Real(0.01, 0.3)
    },
    'CAT': {  # CatBoost
        'iterations': space.Int(20, 100),
        'depth': space.Int(2, 6),
        'learning_rate': space.Real(0.001, 0.1)
    },
    'RF': {  # RandomForest
        'n_estimators': space.Int(10, 100),
        'max_depth': space.Int(2, 6),
        'min_samples_split': space.Int(2, 6),
        'max_features': space.Categorical('sqrt', 'log2', None)
    },
    'KNN': {  # K-Nearest Neighbors with default settings
        'weights': space.Categorical('uniform', 'distance'),  # How neighbors are weighted
        'n_neighbors': space.Int(2, 30), # Number of neighbors
        'p': space.Categorical(1, 2) # Power parameter for the Minkowski metric. When p = 1, this is equivalent to using manhattan_distance (l1), and euclidean_distance (l2) for p = 2
    },
    'XT': { # Extremely randomized trees
        'n_estimators': space.Int(10, 100),
        'max_depth': space.Int(2, 6),
        'min_samples_split': space.Int(2, 6),
        'max_features': space.Categorical('sqrt', 'log2', None)
    }
}

data_sets = {}

# --- Page Configuration ---
st.set_page_config(
    page_title="(AutoGulon) Create a New Experiment",
    page_icon="🤖",
    layout="wide"
)

# back button to return to main page
if st.button('Back'):
    st.switch_page("pages/Training_Models_Options.py")  # Redirect to the main back 

# --- Title ---
st.title("🚀-🤖 Create a New Experiment (AutoGulon)")
st.markdown("Configure and launch a new model training workflow using the advanced AutoGulon framework.")

# Check if client_name was passed
cookies = EncryptedCookieManager(prefix="mlhub_", password="some_secret_key")
if not cookies.ready():
    st.stop()

# Check cookies first
if "client_name" in cookies:
    st.session_state["client_name"] = cookies["client_name"]
    #st.write(cookies["client_name"])
#else:
    #st.error("No found")

# then check in session state
if "client_name" not in st.session_state:
    st.error("No database connection found. Please go back to the main page.")
    st.stop()

client_name = st.session_state["client_name"]

# connect to database
#client = MongoClient('10.14.1.12', 27017)
client = MongoClient(client_name, 27017)

# create the database if it does not already exists
db = client.machine_learning_database
# create tables for models in the databse
models = db.models
# create the results if it does not already exists
results = db.results
# create the results if it does not already exists
datasets = db.datasets
# get all unique exp. names from results collection
#exp_names = db.models.distinct("exp_name", {"type": "AutoGulon"})
exp_names = db.models.distinct("exp_name")
# get all training data names from database
data_names_train = db.datasets.distinct("data_name", {"type": "Train"})
# get all testing data names from database
data_names_list_test = db.datasets.distinct("data_name", {"type": "Test"})

# Conversion function for AutoGluon space to JSON format
def convert_to_json_compatible(hyperparams):
    def serialize(val):
        if isinstance(val, space.Int):
            return {'type': 'int', 'min': val.lower, 'max': val.upper}
        elif isinstance(val, space.Real):
            return {'type': 'real', 'min': val.lower, 'max': val.upper}
        elif isinstance(val, space.Categorical):
            # Try multiple ways to get choices
            if hasattr(val, 'categories'):
                choices = list(val.categories)
            else:
                try:
                    choices = list(val)
                except Exception:
                    raise ValueError(f"Cannot extract choices from Categorical: {val}")
            return {'type': 'categorical', 'choices': choices}
        else:
            raise ValueError(f"Unsupported space type: {val}")

    result = {}
    for model, params in hyperparams.items():
        result[model] = {key: serialize(value) for key, value in params.items()}
    return result

# Conversion from JSON format back to AutoGulon space 
def convert_from_json_compatible(json_hyperparams):
    def deserialize(obj):
        if obj['type'] == 'int':
            return space.Int(obj['min'], obj['max'])
        elif obj['type'] == 'real':
            return space.Real(obj['min'], obj['max'])
        elif obj['type'] == 'categorical':
            return space.Categorical(*obj['choices'])
        else:
            raise ValueError(f"Unknown type: {obj['type']}")

    result = {}
    for model, params in json_hyperparams.items():
        result[model] = {key: deserialize(value) for key, value in params.items()}
    return result

# sample configuration dic
sample_configuration_dic = {
        'time_limit': None, 
        'preset' : 'medium_quality',
        "eval_metric": 'roc_auc',
        "val_set_size": 0.15,
        'keep_only_best': True, 
        'num_bag_folds': None, 
        'num_bag_sets': None, 
        'num_stack_levels': None, 
        "min_postives": 10,
        'cutoff_index': 'youden',
        'custom_hyperparameter_tune_kwargs': {'num_trials': 30, 'scheduler': 'local', 'searcher': 'auto'},
        'custom_hyperparameters': convert_to_json_compatible(custom_hyperparameters_sample)
}


# function to train and generate AutoGulon models
def train_and_generate_models(data_sets, project_name, configuration_dic, unique_value_threshold=10):
    #st.write(configuration_dic)
    #print(num_bag_folds)
    #print(num_stack_levels)    #print(num_bag_setss)
    #print(keep_only_best)
    
    # models dictonary
    models_dictonary = {}

    st.write("---")
    st.subheader("🚀 Starting Experiment...")

    # --- 1. Setup Folders ---
    with st.spinner("Setting up project structure and data..."):
        # get name of the project as part of file name
        project_folder = os.path.join("Models", project_name)
        os.makedirs(project_folder, exist_ok=True)  # Create folder for algorithm results

        train_set = data_sets["Training Set"]
        index_set = data_sets["Index Set"]

        # save the data
        save_data(train_set['Name'], train_set['Data'], os.path.join(project_folder, train_set['Name']))
        save_data(index_set['Name'], index_set['Data'], os.path.join(project_folder, index_set['Name']))

        if train_set['Name'] not in data_names_train:
            st.info(f"Training Dataset {train_set['Name']} is saving in the database", icon="ℹ️")
            # create a list of ML exp.'s that the dataset was used on
            exp_list = [project_name]
            # get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
            # save train set into data folder and database
            os.makedirs("Data Sets", exist_ok=True)
            save_data(train_set['Name'], train_set['Data'], os.path.join("Data Sets", train_set['Name']))
            dataset_train = {
                    "data_name": train_set['Name'],
                    "type": "Train",
                    "time_saved": current_time,
                    "data_path": os.path.join("Data Sets", train_set['Name']),
                    "exps used": exp_list
            }
            datasets.insert_one(dataset_train)
        else:
            st.info(f"Training Dataset {train_set['Name']} of the same name is already in the database", icon="ℹ️")

            # update the dataset in datbase to trackdown the list of ML exps the set was used on
            dataset = datasets.find_one({"data_name": train_set['Name'], "type": "Train"})
            # Get the current list of experiments or initialize it if not present
            exp_list = dataset.get("exps used", [])
            # Add the current project name if it's not already in the list
            if project_name not in exp_list:
                exp_list.append(project_name)

            datasets.update_one(
                {"data_name": train_set['Name'], "type": "Train"}, # Filter condition
                {"$set": { "exps used": exp_list }} # Update operation
            )


        # get proper data name
        if train_set['Name'].endswith('.xlsx'):
            data_name = train_set["Name"][:-len('.xlsx')]
        elif train_set["Name"].endswith('.csv'):
            data_name = train_set["Name"][:-len('.csv')]
        else:
            data_name = train_set["Name"]

        # get input and ouput columns
        input_cols, label_cols, categorical_cols, numeric_cols = parse_exp_multi_outcomes(train_set["Data"], index_set["Data"], unique_value_threshold=unique_value_threshold)
        st.write("Label Columns: ", label_cols)

        # refine the binary outcomes
        df_train = refine_binary_outcomes(train_set["Data"], label_cols)

        #st.write(label_cols)

        for outcome in label_cols:
            with st.spinner(f'Working with outcome: {outcome}'):
                if outcome not in df_train.columns:
                    st.error(f'Unable to generate model because {outcome} is not in dataset.')
                    continue

                # create the train set for the specifc input variables and specific outcome.
                input_train = df_train[input_cols]
                train_data = pd.concat([input_train, df_train[[outcome]]], axis=1)
                #st.write(train_data)

                y_train = train_data[outcome]
                
                # Count positives and negatives
                positives = np.sum(y_train == 1)  # Count instances of 1
                negatives = np.sum(y_train == 0)  # Count instances of 0

                #st.write("Postive Count (on training set):", positives)
                #st.write("Negative Count (on training set):", negatives)

                if positives < configuration_dic['min_postives']:
                    st.error(f"Unable to generate model for {outcome} because of lack of postive outcomes in train set.")
                    continue
                
                # finally start prediction
                st.write('Training has started...')
                predictor = None
                #try:
                predictor = TabularPredictor(label=outcome, eval_metric=configuration_dic['eval_metric']).fit(train_data, 
                                                                                time_limit=configuration_dic['time_limit'], 
                                                                                presets=configuration_dic['preset'],
                                                                                holdout_frac=configuration_dic['val_set_size'],     
                                                                                hyperparameters=configuration_dic['custom_hyperparameters'],  
                                                                                hyperparameter_tune_kwargs=configuration_dic['custom_hyperparameter_tune_kwargs'],
                                                                                num_bag_folds=configuration_dic['num_bag_folds'], 
                                                                                num_stack_levels=configuration_dic['num_stack_levels'], 
                                                                                num_bag_sets=configuration_dic['num_bag_sets'], 
                                                                                raise_on_no_models_fitted=False,
                                                                                keep_only_best=configuration_dic['keep_only_best']) # generate and train a model
                    
                    
                # access the validation results
                results = predictor.fit_summary()
                print(f'Validation Results: {results}.')

                # Create the folder if it doesn't exist
                os.makedirs(os.path.join(project_folder, 'Individual Models'), exist_ok=True)

                # save the model into a joblib file
                file_name = f"{project_folder}/Individual Models/{sanitize_filename(outcome)}_model.joblib"
                joblib.dump(predictor, file_name)
                print(f"Model saved at: {file_name}")
                    
                models_dictonary[outcome] = {}
                models_dictonary[outcome]['Outcome Name'] = outcome
                models_dictonary[outcome]['Model'] = predictor
                models_dictonary[outcome]['Validation Summary'] = results
                #except:
                    #st.error(f"Unable to generate model in {outcome}.")
                    #continue

            
            st.success(f'Training with outcome: {outcome} Done!')
                
    st.success(f'Training is Complete!')

    # save the overall model into a joblib file
    pathway_name = f"{project_folder}/{project_name}_models.joblib"
    joblib.dump(models_dictonary, pathway_name)

    # Finalize Experiment
    st.write("---")
    with st.spinner("Finalizing experiment: saving metadata and final reports..."):
        
        if configuration_dic['custom_hyperparameters'] != None:
            configuration_dic['custom_hyperparameters'] = convert_to_json_compatible(configuration_dic['custom_hyperparameters'])

        # get the current time
        current_datetime = datetime.now()
        current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

        model = {
                "exp_name": project_name,
                "type": "AutoGulon",
                "model_path": pathway_name,
                "input variables": input_cols,
                'outcomes': label_cols,
                "configuration": configuration_dic,
                "train_data": train_set["Name"],
                "time_created": current_time
        }

        models.insert_one(model) # insert one dictonary 

    return models_dictonary

def create_column_summary(df, index_df): # get a summry of the dataframe
    # get the selected columns
    inputs = index_df.columns[index_df.iloc[0] == 1].tolist()
    #st.write(inputs) 

    summary_list = []
    for col in df[inputs].columns:
        col_type = df[col].dtype
        missing_count = df[col].isnull().sum()
        missing_percent = f"{missing_count / len(df) * 100:.2f}%"
        unique_count = df[col].nunique()
        
        col_summary = {
            "Column": col, "Data Type": str(col_type), "Missing Values": missing_count,
            "Missing (%)": missing_percent, "Unique Values": unique_count
        }
        
        if np.issubdtype(col_type, np.number):
            stats = df[col].describe()
            col_summary.update({
                "Mean": f"{stats.get('mean', 0):.2f}", "Std Dev": f"{stats.get('std', 0):.2f}",
                "Min": stats.get('min', 0), "Max": stats.get('max', 0)
            })
        summary_list.append(col_summary)
    return pd.DataFrame(summary_list)

def create_experiment_summary(data_df, index_df): #create a summery of the experiment
    summary_list = []
    for exp_name, row in index_df.iterrows():
        inputs = [col for col, val in row.items() if val == 1]
        outcomes = [col for col, val in row.items() if val == 2]
        
        for outcome in outcomes:
            prevalence = "N/A"
            if outcome in data_df.columns and set(data_df[outcome].dropna().unique()).issubset({0, 1}):
                prev_rate = data_df[outcome].value_counts(normalize=True).get(1, 0)
                prevalence = f"{prev_rate * 100:.2f}%"
            summary_list.append({
                "Experiment Plan": exp_name, "Input Count": len(inputs),
                "Outcome": outcome, "Prevalence": prevalence
            })
    return pd.DataFrame(summary_list)


# gnerate an index file
def gen_idx(df_train, train_set_name, nexp:int=1, sheets:list=None):

    if train_set_name.endswith('.xlsx'):
        filename = train_set_name[:-len('.xlsx')]
    elif train_set_name.endswith('.csv'):
        filename = train_set_name[:-len('.csv')]
    else:
        filename = train_set_name

    indexFileName = filename + '_index.xlsx'
    
    all_columns = list(df_train.columns)

    df_index = pd.DataFrame(columns=all_columns)
    df_index.loc['Key'] = 0

    # convert df_index to an Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_index.to_excel(writer, index=True)

    st.download_button(
        label="Download the Index File Template to set inputs and outputs ⬇️",
        data=output.getvalue(),
        file_name=indexFileName,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return indexFileName

# (The rest of the UI code is the same as before)
# ...
if 'training_method' not in st.session_state:
    #st.session_state.training_method = "Train/Test Split"
    st.session_state.exp_name = ""
    st.session_state.algorithms = []

# --- Step 1: Experiment Setup ---
st.header("Step 1: Define Experiment and Data Strategy")
project_name = st.text_input("Experiment Name", help="Enter a unique name for this experiment.")
#st.radio(
#    "Select Training Method",
#    ["Train/Test Split", "Train Whole Set"],
#    key="training_method",
#    horizontal=True,
#    help="Choose how to evaluate model."
#)

# check if project name already exists in database
if project_name in exp_names:
    st.error("Experiment with that name already exists")
    is_valid = True
    configuration_dic = None
elif project_name is None or project_name=="" or project_name.isspace():
    st.info("Please enter a name for the experiment")
    is_valid = True
    configuration_dic = None
elif '/' in project_name or '\\' in project_name:
    st.error("Invalid Experiment Name (Cannot have / or \\)")
    is_valid = True
    configuration_dic = None
else:
    is_valid = False

# --- Step 2: Upload Data (Dynamic UI) ---
st.header("Step 2: Upload Data")

#user chooses whatever to upload the data or retrive a past data set from the database
data_options = st.radio("Choose an option:", ["Upload a dataset", "Retrive dataset from database"])

if data_options == "Upload a dataset":
    data_name_train = None
    col1, col2 = st.columns(2)
    with col1:
        # Use a consistent key for the main dataset uploader
        #dataset_uploader_key = "main_dataset_uploader"
        
        #if st.session_state.training_method == "Train/Test Split":
        #    main_uploaded_file = st.file_uploader("Upload Dataset (CSV or Excel)", type=['csv', 'xlsx'])
        #else: # Train Whole Set
            #main_uploaded_file = st.file_uploader("Upload Training Dataset", type=['csv', 'xlsx'], key=dataset_uploader_key)
        main_uploaded_file = st.file_uploader("Upload Dataset (CSV or Excel)", type=['csv', 'xlsx'])

    with col2:
        completed_index_file = st.file_uploader("Upload **Completed** Index File", type=['xlsx'])
else:
    main_uploaded_file = None

    # Dropdown to select the training dataset
    data_name_train = st.selectbox("Select a Training Dataset from the database:", data_names_train, index=None, placeholder="Select One...")

    # Open and wrap the file like an uploaded file (binary mode)
    if data_name_train is not None:
        completed_index_file = st.file_uploader("Upload **Completed** Index File", type=['xlsx'])

st.write("")

# upload the dataset(s)

# training sets
if data_options == "Upload a dataset" and main_uploaded_file:
    train_set = main_uploaded_file.name
    #st.write(train_set)

    df_train = load_data(train_set, main_uploaded_file)

    #st.write(df_train.head())  # Display the first few rows

    data_sets["Training Set"] = {}
    data_sets["Training Set"]["Name"] = main_uploaded_file.name
    data_sets["Training Set"]["Data"] = df_train
elif data_name_train:
    df_train = upload_data(os.path.join("Data Sets",data_name_train))
    #st.write(df_train.head())  # Display the first few rows
    data_sets["Training Set"] = {}
    data_sets["Training Set"]["Name"] = data_name_train
    data_sets["Training Set"]["Data"] = df_train
    

if (main_uploaded_file or data_name_train) and not completed_index_file:
    st.info("A dataset has been uploaded. Now generate a matching index file to edit.")
    gen_idx(df_train, data_sets["Training Set"]["Name"])

# Add feedback to the user once they've uploaded their completed file.
elif (main_uploaded_file or data_name_train) and completed_index_file:

    # upload the index file
    index_set = pd.read_excel(completed_index_file)
    #st.write(index_set.head())  # Display the first few rows

    data_sets["Index Set"] = {}
    data_sets["Index Set"]["Name"] = completed_index_file.name
    data_sets["Index Set"]["Data"] = index_set

    #st.success("✅ Completed index file has been uploaded.")
    if st.button("📊 View Data Summary and Experiment Plan", use_container_width=True):
        # When button is clicked, open a dialog
        st.dialog("Data Summary")
        st.header("Experiment Plan Summary")
        st.markdown("Table: summarizes the inputs and outcomes.")
            
        exp_summary_df = create_experiment_summary(df_train, index_set)
        st.dataframe(exp_summary_df, use_container_width=True)

        st.header("Dataset Column Analysis")
        st.markdown("Table: Statistics for each column in the dataset.")
        col_summary_df = create_column_summary(df_train, index_set)
        st.dataframe(col_summary_df, use_container_width=True)

        if st.button("Close"):
            st.rerun() # Closes the dialog

# --- Step 3: Configure Training Pipeline ---
st.header("Step 3: Configure Training Pipeline")

if (main_uploaded_file or data_name_train) and completed_index_file and is_valid==False:
    configuration_dic = None
    # choose how you configure your model(s)
    configure_options = st.radio("Choose an option:", ["Upload a file", "User Customization"], help="User can either use the UI customization to set up the experiment or they can generate a configuration file where they can also set up a hyperparameter feature space.")
else:
    configure_options = None
    configuration_dic = None

# if the user chooses UI customization
if configure_options == "User Customization":

    # min. number of postive cases required for an outcome to be trained on
    unique_value_threshold = st.number_input("Enter the minimum unique value threshold for a input variable to be consired catagorical:", min_value=1, max_value=100, value=10, help="Min. number of postive cases required for an outcome to be trained on.")

    # cutoff type for binary classification
    threshold_type = st.selectbox("Optimal Cutoff Threshold Method", ['youden', 'mcc', 'ji', 'f1'], help="Cutoff type for binary classification.")

    # user setup for parameters for the AutoGulon model
    with st.expander("▶️ AutoGulon Parameters"):

        # min. number of postive cases required to train a model for an outcome
        min_postives = st.number_input("Minimum Positive Cases for an Outcome", min_value=1, value=10)

        # set the time limit
        is_time_limit = st.toggle("Time Limit?")
        if is_time_limit:
            time_limit = st.number_input("Time Limit", min_value=100, value=3500, help="Time limit for model training.")
        else:
            time_limit = None

        # select the quality model
        preset = st.selectbox("Presets", ['best_quality', 'high_quality', 'good_quality', 'medium_quality', 'experimental_quality', 'optimize_for_deployment', 'interpretable', 'ignore_text'], help="Preset configurations for various arguments in fit(). Can significantly impact the quality of model and predictive accuracy.")
        
        # select valuation metric
        eval_metric = st.selectbox("Evaluation metric", ['roc_auc', 'accuracy', 'balanced_accuracy', 'f1', 'f1_macro', 'log_loss', 'precision', 'recall'], help="Metric by which predictions will be ultimately evaluated on test data.")

        # set the validation size
        val_set_size = st.slider("Validation Set Size", 0.1, 0.8, 0.15, help="Fraction of train_data to holdout as tuning data for optimizing hyperparameters.")

        # keep best models
        keep_only_best = st.radio("Keep Only Best Models?", ("True", "False"), index=1, horizontal=True, help="If True, only the best model and its ancestor models are saved in the outputted predictor. All other models are deleted.")

        # set the number of bagging folds
        bag_folds = st.toggle("Bag Folds?")
        if bag_folds:
            num_bag_folds = st.number_input("Number of bagging folds", min_value=2, value=5, help="Number of folds used for bagging of models. When num_bag_folds = k, training time is roughly increased by a factor of k.")
        else:
            num_bag_folds = None

        # Set the number of bagging sets
        bag_sets = st.toggle("Bag Sets?")
        if bag_sets:
            num_bag_sets = st.number_input("Number of bagging sets", min_value=2, value=5, help="Number of repeats of kfold bagging to perform.")
        else:
            num_bag_sets = None

        # set the number of stack levels
        stack_levels = st.toggle("Stack Levels?")
        if stack_levels:
            num_stack_levels = st.number_input("Number of stacking levels", min_value=2, value=5, help="Number of stacking levels to use in stack ensemble. Roughly increases model training time by factor of num_stack_levels+1.")
        else:
            num_stack_levels = None

    # Define hyperparameter tuning settings
    with st.expander("▶️ Hyperparameter Tuning Setting"):

        # should we do hyperparmeter tuning?
        hyperparameter_tune = st.toggle("Hyperparameter Tuning?")
        if hyperparameter_tune:
            # number of trials
            num_trials = st.number_input("Number of trials", min_value=1, value=30, help="How many HPO trials to run ‘scheduler’.")
    
            # Which scheduler algorithm to use
            scheduler = st.selectbox("Scheduler Strategy", ['local', 'hyperband', 'fifo', 'bayesopt'], help="The scheduler in hyperparameter_tune_kwargs controls how trials (different hyperparameter configurations) are managed and evaluated during hyperparameter tuning.")

            # Which searching algorithm or ’searcher’ to use
            searcher = st.selectbox("Searching Strategy", ['auto', 'random', 'bayesopt', 'grid'], help="The searcher in AutoGluon's hyperparameter_tune_kwargs specifies how hyperparameter configurations are selected during hyperparameter tuning.")

            custom_hyperparameter_tune_kwargs = {'num_trials': num_trials, 'scheduler': scheduler, 'searcher': searcher}

            # File uploader for file for the feature space
            uploaded_custom_hyperparameters= st.file_uploader("Upload a Hyperparameter Feature Space")

            if uploaded_custom_hyperparameters:
                custom_hyperparameters_json = json.load(uploaded_custom_hyperparameters)
                custom_hyperparameters = convert_from_json_compatible(custom_hyperparameters_json)
                st.write(custom_hyperparameters)
            else:
                custom_hyperparameters = None

        else:
            custom_hyperparameter_tune_kwargs = None
            custom_hyperparameters = None


        # download a template file for the feature space
        st.download_button(
            label="Download the Sample File for Hyperparameter Feature Space ⬇️",
            data=json.dumps(convert_to_json_compatible(custom_hyperparameters_sample), indent=2),
            file_name=f"{project_name}_feature_space.json",
            mime="application/json"
        )

    configuration_dic = {
        'time_limit': time_limit, 
        'preset' : preset,
        "eval_metric": eval_metric,
        "val_set_size": val_set_size,
        'keep_only_best': keep_only_best, 
        'num_bag_folds': num_bag_folds, 
        'num_bag_sets': num_bag_sets, 
        'num_stack_levels': num_stack_levels, 
        "min_postives": min_postives,
        'cutoff_index': threshold_type,
        'custom_hyperparameter_tune_kwargs': custom_hyperparameter_tune_kwargs,
        'custom_hyperparameters': custom_hyperparameters
    }

elif configure_options == "Upload a file":
    unique_value_threshold = st.number_input("Enter the minimum unique value threshold for a input variable to be consired catagorical:", min_value=1, max_value=100, value=10)
    # Number input
    # reset configuration_dic
    configuration_dic = None
    # sends user to a seperate page where they can download an empty configuration file where they can customize themselves
    st.download_button(
        label="Download Configuration File Template ⬇️",
        data=json.dumps(sample_configuration_dic, indent=2),
        file_name=f"{project_name}.json",
        mime="application/json"
    )

    with st.expander("▶️ List of AutoGulon Algorithims to use."):
        st.write("Please use the short versions of the algo names shown on the **right**.")
        st.write(algo_shortnames)

    # File uploader for the training
    uploaded_config_file= st.file_uploader("Upload a Configuration File")

    if uploaded_config_file:
        configuration_dic = json.load(uploaded_config_file)
        custom_hyperparameters = configuration_dic['custom_hyperparameters']
        if custom_hyperparameters != None:
            configuration_dic['custom_hyperparameters'] = convert_from_json_compatible(configuration_dic['custom_hyperparameters'])
        st.write(configuration_dic)

else:
    configuration_dic = None

# --- Step 4: Execute ---
st.header("Step 4: Run Experiment")

if configuration_dic != None and st.button("Start Training", type="primary", use_container_width=True):
        
    # Call the main orchestration function
    models_dictonary = train_and_generate_models(data_sets, project_name, configuration_dic, unique_value_threshold=unique_value_threshold)

    #st.write(models_dictonary)

    st.subheader("Jump to Testing the Models") # redirect to the testing section
    st.page_link("pages/Testing_AutoGulon_Models.py", label="Test Models", icon="🧪")
