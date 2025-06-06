import time
import os
import pandas as pd
import numpy as np
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn import metrics
import time
import psutil
import autogluon
import pandas as pd
import numpy as np
import json
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from scipy import interp
from scipy.stats import norm
from scipy.special import ndtri
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import magic 
import pickle
import random 
from random import randint
from random import uniform
from scipy import stats
import ast
from pathlib import Path
import joblib as joblib
from joblib import dump, load
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from sklearn.model_selection import cross_validate
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import RocCurveDisplay
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.metrics import auc
from sklearn.utils import resample
from sklearn.model_selection import KFold
from sklearn.model_selection import StratifiedKFold, RepeatedStratifiedKFold
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import VarianceThreshold
from sklearn.model_selection import RandomizedSearchCV, GridSearchCV
#from tune_sklearn import TuneSearchCV
from skopt import BayesSearchCV
from sklearn.calibration import CalibratedClassifierCV
from sklearn.calibration import calibration_curve, CalibrationDisplay
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score
from sklearn.metrics import classification_report
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import label_binarize
from sklearn.metrics import confusion_matrix
from scipy import interp
from scipy.stats import norm
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
import random 
from random import randint
from random import uniform
from scipy import stats
import mrmr
from mrmr import mrmr_classif
import xgboost
import catboost
import shap
from scipy import stats
import os
import joblib as joblib
from joblib import dump, load
import json
import tkinter as tk
from tkinter import *
from autogluon.tabular import TabularDataset, TabularPredictor
#np.random.seed(1000)
from difflib import SequenceMatcher
#np.random.seed(1000)
rstate = 12

# import module
import datetime
import pprint
import pymongo
from pymongo import MongoClient
import streamlit as st

global options_default
global options_test_set

options_default = { # default options dict
        'oneHotEncode' : "True", 
        'Impute': "True", 
        'cutMissingRows' : "True",
        "cut threshold": 0.60,
        "inf": 'replace with null',
        'outliers': "log",
        'outliers_N': 50000, 
        'Scaling': "True", 
        'scalingMethod': "StandardScaler", 
        'QuantileTransformer': "True", 
        'Normalize': "True",
        'rebalance' : "True",
        'rebalance_type': "SMOTE",
        'FeatureSelection': "True",
        "method": 'SelectKBest-f_classif',
        'N_features': 40, 
        'strategy': "random", 
        'itr': 50,
        "CV": 10,
        "n_repeats": 1,
        "min_postives": 10,
    }

options_test_set = { # test set options dict
        'oneHotEncode' : "True", 
        'Impute': "True", 
        'cutMissingRows' : "True",
        "cut threshold": 0.60,
        "inf": 'replace with null',
        'outliers': "log",
        'outliers_N': 50000, 
        'Scaling': "True", 
        'scalingMethod': "StandardScaler", 
        'QuantileTransformer': "True", 
        'Normalize': "True",
        'rebalance' : "True",
        'rebalance_type': "SMOTE",
        'FeatureSelection': "True",
        "method": 'SelectKBest-f_classif',
        'N_features': 40, 
        'strategy': "random", 
        'itr': 50,
        "CV": 10,
        "n_repeats": 1,
        "min_postives": 10,
        "test_size": 0.2,
        'cutoff_index': 'youden'
    }

def generate_configuration_file(num_exp, project_name, train_set, test_sets, exp_name, algorithms, exp_type, options, param_vals):
    configuration_dic = {}
    configuration_dic[project_name] = {}
    #configuration_dic[project_name]['train_set']  = train_set
    #configuration_dic[project_name]['test_sets']  = test_sets
    configuration_dic[project_name]['exp_type']  = exp_type
    
    for i in range(num_exp):
        print(i)
        experiment = {}
        experiment['algorithm'] = algorithms[i]
        #experiment['training_type'] = training_type[i]
        experiment['options'] = options[i]
        experiment['param_vals'] = param_vals[i]
        
        configuration_dic[project_name][exp_name[i]] = experiment
    
    #st.write("Configuration_dic : ", configuration_dic)
    
    return configuration_dic
    

def generate_configuration_template(project_name, num_exp, test_set):
    print("Number of exp. :", num_exp)
    exp_name=["exp_" + str(i) for i in range(num_exp)]
    algorithms=["enter_algo_here"]*num_exp
    train_set="enter_filename_here"
    options=[options_default]*num_exp if test_set=="Yes" else [options_test_set]*num_exp
    
    configuration_dic = generate_configuration_file(num_exp=num_exp, 
                            project_name=project_name, 
                            train_set=train_set,
                            test_sets=["None"],
                            exp_name=exp_name,
                            algorithms=algorithms, 
                            exp_type="enter_type_here", 
                            #training_type=["enter_type_here"] * num_exp,
                            options=options,
                            param_vals=["None"]*num_exp)

    #st.write(configuration_dic)

    return configuration_dic

# Title
st.title("Download a Configuration Template File")

st.write("This page allows you to download a Configuration template file with a user-specified number of experiments/models.")

# enter experiment name
exp_name = st.text_input("Enter Name of the ML Experiment", "exp_name")

# Test Set
test_set = st.radio("Do you have a seperate test set(s):", ["Yes", "No"])

# Number input
num_models = st.number_input("Enter the number of models:", min_value=1, max_value=100, value=5)

#configuration_dic = generate_configuration_template(exp_name, num_models)
    
# Convert Python dict to JSON string
#json_configuration_dic = json.dumps(generate_configuration_template(exp_name, num_models), indent=4)

st.download_button(
        label="Download Configuration File Template",
        data=json.dumps(generate_configuration_template(exp_name, num_models, test_set), indent=4),
        file_name=f"{exp_name}.json",
        mime="application/json"
)

st.write()

# back button to return to Training_Native_Models.py
if st.button('Back'):
    path_back = "pages/Training_Native_Models_W_TS.py" if test_set=="Yes" else "pages/Training_Native_Models_WO_TS.py"
    st.switch_page(path_back)  # Redirect to the Training_Native_Models.py

