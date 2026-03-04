from pathlib import Path
import time
import io
import pandas as pd
import numpy as np
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn import metrics
import time
import psutil
import autogluon
import pandas as pd
import numpy as np
import json
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
#from scipy import interp
from scipy.stats import norm
from scipy.special import ndtri
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import magic 
import pickle
import random 
from random import randint
from random import uniform
from scipy import stats
import ast
from pathlib import Path
import joblib as joblib
from joblib import dump, load
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from sklearn.model_selection import cross_validate
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import RocCurveDisplay
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.metrics import auc
from sklearn.utils import resample
from sklearn.model_selection import KFold
from sklearn.model_selection import StratifiedKFold, RepeatedStratifiedKFold
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import VarianceThreshold
from sklearn.model_selection import RandomizedSearchCV, GridSearchCV
#from tune_sklearn import TuneSearchCV
from skopt import BayesSearchCV
from sklearn.calibration import CalibratedClassifierCV
from sklearn.calibration import calibration_curve, CalibrationDisplay
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score
from sklearn.metrics import classification_report
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import label_binarize
from sklearn.metrics import confusion_matrix
#from scipy import interp
from scipy.stats import norm
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
import random 
from random import randint
from random import uniform
from scipy import stats
import mrmr
from mrmr import mrmr_classif
import xgboost
import catboost
import shap
from scipy import stats
import os
import sys
import joblib as joblib
from joblib import dump, load
import json
import tkinter as tk
from tkinter import *
from autogluon.tabular import TabularDataset, TabularPredictor
from autogluon.common import space
from autogluon.common.space import Int, Real, Categorical
from difflib import SequenceMatcher
#np.random.seed(1000)
rstate = 12
# import module
from datetime import datetime, timedelta
import pprint
import pymongo
from pymongo import MongoClient
import signal
import sys
import traceback
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures import ThreadPoolExecutor
from joblib import Parallel, delayed
from multiprocessing import Process
#import fastapi
from fastapi import FastAPI
from celery import Celery
import uuid
import threading
from Multi_Outcome_Classification_tools import multi_outcome_hyperparameter_binary, multi_outcome_hyperparameter_binary_train_and_test, multi_outcome_cv
from Common_Tools import sanitize_filename, convert_to_json_compatible, generate_configuration_file, generate_configuration_template, generate_results_table, generate_congfig_file, get_avg_results_dic, wrap_text_excel, expand_cell_excel, grid_excel, generate_all_idx_files, upload_data, load_data, save_data, data_prep, data_prep_train_set, parse_exp_multi_outcomes, setup_multioutcome_binary, refine_binary_outcomes, generate_joblib_model
from roctools import full_roc_curve, plot_roc_curve

# function to train and generate AutoGulon models
def train_and_generate_models(job_id, data_sets, project_name, configuration_dic, client_name, unique_value_threshold):
    # Create a status file
    #os.makedirs("jobs", exist_ok=True)
    #status_path = f"jobs/{job_id}.json"

    print("Training the models have started!")

    # connect to database
    #client = MongoClient('10.14.1.12', 27017)
    client = MongoClient(client_name, 27017)

    # create the database if it does not already exists
    db = client.machine_learning_database
    # create tables for models in the databse
    models = db.models
    # create the results if it does not already exists
    results = db.results
    # create the results if it does not already exists
    datasets = db.datasets
    # create the jobs if it does not already exists
    jobs = db.jobs
    # get all unique exp. names from results collection
    #exp_names = db.models.distinct("exp_name", {"type": "AutoGulon"})
    exp_names = db.models.distinct("exp_name")
    # get all training data names from database
    data_names_train = db.datasets.distinct("data_name", {"type": "Train"})
    # get all testing data names from database
    data_names_list_test = db.datasets.distinct("data_name", {"type": "Test"})

    # create a jobs dictonary to store in the MongoDB database
    job_doc = {
        "job_id": job_id,
        "job_type": "Training (AutoGulon)",
        "exp_name": project_name,
        "status": "queued",
        "message": "Waiting to start...",
        "created_at": datetime.now(),
        "last_updated": datetime.now()
    }
    db.jobs.insert_one(job_doc)

    def update_status(state, message): # inner function to update the status of the current job
        db.jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": state,
                "message": message,
                "last_updated": datetime.now()
            }})
        
    def update_status_when_done(state, message): # inner function to update the status when it is finsihed
        completed_time = datetime.now()
        expires_at = completed_time + timedelta(days=1)
        db.jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": state,
                "message": message,
                "last_updated": completed_time,
                "completed_at": completed_time,
                "expires_at": expires_at
            }})
        
    def heartbeat():
        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "heartbeat": datetime.now(),
                "last_updated": datetime.now()
            }}
        )

    stop_heartbeat = False
    def heartbeat_loop():
        while not stop_heartbeat:
            heartbeat()
            time.sleep(30)

    heartbeat_thread = threading.Thread(
        target=heartbeat_loop,
        daemon=True
    )
    heartbeat_thread.start()
        
    
    def handle_termination(signum, frame):
        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "Stopped",
                "message": "Job was terminated unexpectedly.",
                "last_updated": datetime.now(),
                "expires_at": datetime.now() + timedelta(days=1)
            }}
        )
        sys.exit(1)

    signal.signal(signal.SIGTERM, handle_termination)
    signal.signal(signal.SIGINT, handle_termination)

    try:
        print("Starting Experiment...")

        # models dictonary
        models_dictonary = {}

        # dictonary to show excluded columns
        removed_label_cols = {}

        # --- 1. Setup Folders ---
        # get name of the project as part of file name
        project_folder = os.path.join("Models", project_name)
        os.makedirs(project_folder, exist_ok=True)  # Create folder for algorithm results

        train_set = data_sets["Training Set"]
        index_set = data_sets["Index Set"]
        
        update_status("running", "Saving the training data...")

        # save the data
        #save_data(train_set['Name'], train_set['Data'], os.path.join(project_folder, train_set['Name']))
        save_data(index_set['Name'], index_set['Data'], os.path.join(project_folder, index_set['Name']))

        if train_set['Name'] not in data_names_train:
            # create a list of ML exp.'s that the dataset was used on
            exp_list = [project_name]
            # get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
            # save train set into data folder and database
            os.makedirs("Data Sets", exist_ok=True)
            save_data(train_set['Name'], train_set['Data'], os.path.join("Data Sets", train_set['Name']))
            dataset_train = {
                    "data_name": train_set['Name'],
                    "type": "Train",
                    "time_saved": current_time,
                    "data_path": os.path.join("Data Sets", train_set['Name']),
                    "exps used": exp_list
            }
            datasets.insert_one(dataset_train)
        else:
            save_data(train_set['Name'], train_set['Data'], os.path.join("Data Sets", train_set['Name']))
            # update the dataset in datbase to trackdown the list of ML exps the set was used on
            dataset = datasets.find_one({"data_name": train_set['Name'], "type": "Train"})
            # Get the current list of experiments or initialize it if not present
            exp_list = dataset.get("exps used", [])
            # Add the current project name if it's not already in the list
            if project_name not in exp_list:
                exp_list.append(project_name)

            datasets.update_one(
                {"data_name": train_set['Name'], "type": "Train"}, # Filter condition
                {"$set": { "exps used": exp_list }} # Update operation
            )


            # get proper data name
            if train_set['Name'].endswith('.xlsx'):
                data_name = train_set["Name"][:-len('.xlsx')]
            elif train_set["Name"].endswith('.csv'):
                data_name = train_set["Name"][:-len('.csv')]
            else:
                data_name = train_set["Name"]

            # get input and ouput columns
            input_cols, label_cols, categorical_cols, numeric_cols = parse_exp_multi_outcomes(train_set["Data"], index_set["Data"], unique_value_threshold=unique_value_threshold)

            # refine the binary outcomes
            df_train = refine_binary_outcomes(train_set["Data"], label_cols)

            #st.write(label_cols)

        # --- 2. Training Models ---
        update_status("running", "Starting training...")

        t2 = time.time()

        for outcome in label_cols:

            if outcome not in df_train.columns:
                print(f'Unable to generate model because {outcome} is not in dataset.')
                continue

            # create the train set for the specifc input variables and specific outcome.
            input_train = df_train[input_cols]
            train_data = pd.concat([input_train, df_train[[outcome]]], axis=1)
            #st.write(train_data)

            y_train = train_data[outcome]
            
            # Count positives and negatives
            positives = np.sum(y_train == 1)  # Count instances of 1
            negatives = np.sum(y_train == 0)  # Count instances of 0

            #st.write("Postive Count (on training set):", positives)
            #st.write("Negative Count (on training set):", negatives)

            if positives < configuration_dic['min_postives']:
                print(f"Unable to generate model for {outcome} because of lack of postive outcomes in train set.")
                removed_label_cols[outcome] = positives
                continue
            
            # finally start prediction
            print('Training has started...')
            update_status("running", f"Training with {outcome}...")
            predictor = None
            #try:
            predictor = TabularPredictor(label=outcome, eval_metric=configuration_dic['eval_metric']).fit(train_data, 
                                                                            time_limit=configuration_dic['time_limit'], 
                                                                            presets=configuration_dic['preset'],
                                                                            holdout_frac=configuration_dic['val_set_size'],     
                                                                            hyperparameters=configuration_dic['custom_hyperparameters'],  
                                                                            hyperparameter_tune_kwargs=configuration_dic['custom_hyperparameter_tune_kwargs'],
                                                                            num_bag_folds=configuration_dic['num_bag_folds'], 
                                                                            num_stack_levels=configuration_dic['num_stack_levels'], 
                                                                            num_bag_sets=configuration_dic['num_bag_sets'], 
                                                                            raise_on_no_models_fitted=False,
                                                                            keep_only_best=configuration_dic['keep_only_best']) # generate and train a model
                
                
            # access the validation results
            fit_summary = predictor.fit_summary()
            print(f'Validation Results: {fit_summary}.')

            # Create the folder if it doesn't exist
            os.makedirs(os.path.join(project_folder, 'Individual Models'), exist_ok=True)

            # save the model into a joblib file
            file_name = f"{project_folder}/Individual Models/{sanitize_filename(outcome)}_model.joblib"
            joblib.dump(predictor, file_name)
            print(f"Model saved at: {file_name}")
                
            models_dictonary[outcome] = {}
            models_dictonary[outcome]['Outcome Name'] = outcome
            models_dictonary[outcome]['Model'] = predictor
            models_dictonary[outcome]['Validation Summary'] = fit_summary

            
            print(f'Training with outcome: {outcome} Done!')
            
        t1 = time.time() 
        print(f'The Time Training Took :{t1-t2} seconds')

        # save a textfile telling the time it took to test models
        filename_time = os.path.join(project_folder, "testing_time.txt")
        f = open(filename_time, "w", encoding="utf-8")
        f.write(f'The Time Training Took :{t1-t2} seconds')
        f.close() 
                
        print(f'Training is Complete!')

        # save the overall model into a joblib file
        pathway_name = f"{project_folder}/{project_name}_models.joblib"
        joblib.dump(models_dictonary, pathway_name)

        with open(os.path.join(project_folder, "excluded_label_cols_setup.txt"), "w", encoding="utf-8") as file:
            file.write('Label Columns not included due to too little postive labels (outcome : num postives):  %s' % json.dumps(removed_label_cols, indent=4))
            file.write('Postive Labels cutoff is:  %s' % configuration_dic['min_postives'])
        file.close()

        # --- 3. Finalize Experiment
            
        if configuration_dic['custom_hyperparameters'] != None:
            configuration_dic['custom_hyperparameters'] = convert_to_json_compatible(configuration_dic['custom_hyperparameters'])

        # get the current time
        current_datetime = datetime.now()
        current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

        model = {
                "exp_name": project_name,
                "type": "AutoGulon",
                "model_path": pathway_name,
                "input variables": input_cols,
                'outcomes': label_cols,
                "configuration": configuration_dic,
                "train_data": train_set["Name"],
                "time_created": current_time
        }

        models.insert_one(model) # insert one dictonary 

        update_status_when_done("Completed", "Training complete")

        return models_dictonary
    
    except Exception as e:
        error_message = str(e)
        stack_trace = traceback.format_exc()

        print(f"Error Occured. Experiment '{project_name}' failed!")
        traceback.print_exc()

        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "Failed",
                "message": error_message,
                "error_trace": stack_trace,
                "last_updated": datetime.now(),
                "expires_at": datetime.now() + timedelta(days=1)
            }}
        )

    finally:
        stop_heartbeat = True
        heartbeat_thread.join(timeout=1)