from pathlib import Path
import time
import io
import pandas as pd
import numpy as np
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn import metrics
import time
import psutil
import autogluon
import pandas as pd
import numpy as np
import json
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
#from scipy import interp
from scipy.stats import norm
from scipy.special import ndtri
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import magic 
import pickle
import random 
from random import randint
from random import uniform
from scipy import stats
import ast
from pathlib import Path
import joblib as joblib
from joblib import dump, load
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from sklearn.model_selection import cross_validate
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import RocCurveDisplay
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.metrics import auc
from sklearn.utils import resample
from sklearn.model_selection import KFold
from sklearn.model_selection import StratifiedKFold, RepeatedStratifiedKFold
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import VarianceThreshold
from sklearn.model_selection import RandomizedSearchCV, GridSearchCV
#from tune_sklearn import TuneSearchCV
from skopt import BayesSearchCV
from sklearn.calibration import CalibratedClassifierCV
from sklearn.calibration import calibration_curve, CalibrationDisplay
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score
from sklearn.metrics import classification_report
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import label_binarize
from sklearn.metrics import confusion_matrix
#from scipy import interp
from scipy.stats import norm
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
import random 
from random import randint
from random import uniform
from scipy import stats
import mrmr
from mrmr import mrmr_classif
import xgboost
import catboost
import shap
from scipy import stats
import os
import signal
import sys
import traceback
import sys
import joblib as joblib
from joblib import dump, load
import json
import tkinter as tk
from tkinter import *
from autogluon.tabular import TabularDataset, TabularPredictor
from autogluon.common import space
from autogluon.common.space import Int, Real, Categorical
from difflib import SequenceMatcher
#np.random.seed(1000)
rstate = 12
# import module
from datetime import datetime, timedelta
import pprint
import pymongo
from pymongo import MongoClient
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures import ThreadPoolExecutor
from joblib import Parallel, delayed
from multiprocessing import Process
#import fastapi
from fastapi import FastAPI
from celery import Celery
import uuid
import threading
from Multi_Outcome_Classification_tools import multi_outcome_hyperparameter_binary, multi_outcome_hyperparameter_binary_train_and_test, multi_outcome_cv
from Common_Tools import sanitize_filename, convert_to_json_compatible, generate_configuration_file, generate_configuration_template, generate_results_table, generate_congfig_file, get_avg_results_dic, wrap_text_excel, expand_cell_excel, grid_excel, generate_all_idx_files, upload_data, load_data, save_data, data_prep, data_prep_train_set, parse_exp_multi_outcomes, setup_multioutcome_binary, refine_binary_outcomes, generate_joblib_model
from roctools import full_roc_curve, plot_roc_curve

def plot_feature_importance(exp_name, feature_importance_dic, directory_name):
    outcomes = list(feature_importance_dic.keys())

    for outcome in outcomes:
        # get the table
        df_feature_importance = feature_importance_dic[outcome]
        
        # the abs value of feature importance values
        df_feature_importance['importance (abs)'] = df_feature_importance['importance'].abs()
        
        # Sort the DataFrame by the column values
        #df_feature_importance = df_feature_importance.sort_values(by='importance (abs)', ascending=False)

        # get the names of the top 10 most important features
        top_features = df_feature_importance['feature'].head(10).tolist()
        top_importances = df_feature_importance['importance (abs)'].head(10).tolist()
        
        # plot the 10 top most important features
        fig, ax = plt.subplots(figsize=(6, 6))
        #ax = df_feature_importance['importance (abs)'][:10].plot(kind='bar')
        ax.bar(top_features, top_importances, color='skyblue')
        ax.set_title(f'Feature Importance for {outcome} on {exp_name} (Top 10)', fontsize=8)
        ax.set_ylabel('Importance (abs)', fontsize=7)
        ax.set_xlabel('Feature', fontsize=7)

        # Add text labels
        for i, v in enumerate(top_importances):
            ax.text(i, v + 0.0005, f'{v:.3f}', ha='center')

        # Set y-axis range from 0 to 1
        ax.set_ylim(0, max(top_importances) + 0.01)

        # Adjust x-axis labels for readability
        plt.xticks(top_features, rotation=45, ha='right', fontsize=8)  # Rotate and align right
        plt.yticks(fontsize=8)
        
        # Save the plot as a PNG file
        os.makedirs(directory_name, exist_ok=True)
        file_name = f'{directory_name}/{sanitize_filename(outcome)}_feature_importance.png'
        plt.savefig(file_name, dpi=300, bbox_inches='tight')

        plt.close()


def generate_results_table(results_dictonary, outcomes):
    # Initialize an empty list to collect rows
    rows = []
    no_results = []

    for outcome in outcomes:
        print(outcome)
        try:
            print(results_dictonary[outcome])
        except:
            print(f'No results exists for {outcome}.')
            no_results.append(outcome)
            continue
            
        leaderboard = results_dictonary[outcome]['leaderboard']
        df_leaderboard = pd.DataFrame(leaderboard)
        best_model = df_leaderboard.loc[0]['model']
        best_model_auroc = df_leaderboard.loc[0]['score_test']

        if 'AUROC Score (Train)' in list(results_dictonary[outcome]['evaluation'].keys()):
            new_row = {'Outcome': outcome,
                  'AUROC Score': results_dictonary[outcome]['evaluation']['AUROC Score'],
                  'AUROC CI Lower': results_dictonary[outcome]['evaluation']['AUROC CI Low'],
                  'AUROC CI Upper': results_dictonary[outcome]['evaluation']['AUROC CI High'],
                  'Accuracy': results_dictonary[outcome]['evaluation']['Test Accuracy'],
                  'Precision': results_dictonary[outcome]['evaluation']['precision'],
                  'Recall': results_dictonary[outcome]['evaluation']['recall'],
                  'F1 Score': results_dictonary[outcome]['evaluation']['f1 score'],
                  'TPR': results_dictonary[outcome]['evaluation']['TPR'], # same as Sensitivity 
                  'TNR': results_dictonary[outcome]['evaluation']['TNR'], # same as Specificity 
                  'FPR': results_dictonary[outcome]['evaluation']['FPR'], 
                  'FNR': results_dictonary[outcome]['evaluation']['FNR'],
                  'PPV': results_dictonary[outcome]['evaluation']['PPV'],
                  'NPV': results_dictonary[outcome]['evaluation']['NPV'],
                  'TP': results_dictonary[outcome]['evaluation']['TP'],
                  'FP': results_dictonary[outcome]['evaluation']['FP'],
                  'TN': results_dictonary[outcome]['evaluation']['TN'],
                  'FN': results_dictonary[outcome]['evaluation']['FN'],
                  'Cutoff value': results_dictonary[outcome]['evaluation']['cutoff'],
                  'Best Model': best_model,
                  'Best Model AUROC Score': best_model_auroc,
                  'P': results_dictonary[outcome]['evaluation']['P'],
                  'N': results_dictonary[outcome]['evaluation']['N'],
                  'AUROC Score (Train)': results_dictonary[outcome]['evaluation']['AUROC Score (Train)'],
                  'AUROC CI Lower (Train)': results_dictonary[outcome]['evaluation']['AUROC CI Low (Train)'],
                  'AUROC CI Upper (Train)': results_dictonary[outcome]['evaluation']['AUROC CI High (Train)'],
                  'P (Train)': results_dictonary[outcome]['evaluation']['P (Train)'],
                  'N (Train)': results_dictonary[outcome]['evaluation']['N (Train)']}
        else:
            new_row = {'Outcome': outcome,
                  'AUROC Score': results_dictonary[outcome]['evaluation']['AUROC Score'],
                  'AUROC CI Lower': results_dictonary[outcome]['evaluation']['AUROC CI Low'],
                  'AUROC CI Upper': results_dictonary[outcome]['evaluation']['AUROC CI High'],
                  'Accuracy': results_dictonary[outcome]['evaluation']['Test Accuracy'],
                  'Precision': results_dictonary[outcome]['evaluation']['precision'],
                  'Recall': results_dictonary[outcome]['evaluation']['recall'],
                  'F1 Score': results_dictonary[outcome]['evaluation']['f1 score'],
                  'TPR': results_dictonary[outcome]['evaluation']['TPR'], # same as Sensitivity 
                  'TNR': results_dictonary[outcome]['evaluation']['TNR'], # same as Specificity 
                  'FPR': results_dictonary[outcome]['evaluation']['FPR'], 
                  'FNR': results_dictonary[outcome]['evaluation']['FNR'],
                  'PPV': results_dictonary[outcome]['evaluation']['PPV'],
                  'NPV': results_dictonary[outcome]['evaluation']['NPV'],
                  'TP': results_dictonary[outcome]['evaluation']['TP'],
                  'FP': results_dictonary[outcome]['evaluation']['FP'],
                  'TN': results_dictonary[outcome]['evaluation']['TN'],
                  'FN': results_dictonary[outcome]['evaluation']['FN'],
                  'Cutoff value': results_dictonary[outcome]['evaluation']['cutoff'],
                  'Best Model': best_model,
                  'Best Model AUROC Score': best_model_auroc,
                  'P': results_dictonary[outcome]['evaluation']['P'],
                  'N': results_dictonary[outcome]['evaluation']['N']}
            
        # Optionally add training confusion matrix values if available
        train_metrics = ['TP (Train)', 'FP (Train)', 'TN (Train)', 'FN (Train)']

        for metric in train_metrics: # add training confusion matrix numbers if they exist. (Train-Test)/Train %
            if metric in list(results_dictonary[outcome]['evaluation'].keys()):
                new_row[metric] = results_dictonary[outcome]['evaluation'][metric]

        # calcuate and add the % change of the train vs. test AUROC score
        if 'AUROC Score (Train)' in list(results_dictonary[outcome]['evaluation'].keys()):
            new_row['Train vs. Test AUROC change%'] = ((results_dictonary[outcome]['evaluation']['AUROC Score (Train)'] - results_dictonary[outcome]['evaluation']['AUROC Score']) / results_dictonary[outcome]['evaluation']['AUROC Score (Train)']) * 100
            
        # add new row
        rows.append(new_row)
        
    # Specifying the list of outcomes with no results for demonstration purposes
    print(f'Outcomes with no results are {no_results}')

    # Writing the output to a text file
    with open("outcomes_no_results.txt", "w") as file:
        file.write(f"Outcomes with no results are {no_results}")
            
    # Convert the list of rows into a DataFrame
    results_df = pd.DataFrame(rows)
    return results_df
        

# testing models function
def test_model(models, test_data_raw, input_columns, outcomes, train_data_raw=None, cutoff_index='youden', algorithm_folder="Logfiles"):

    # Write the starting message
    #placeholder.write("Currently testing the models...")
    print("Currently testing the models...")

    print(f"Cutoff: {cutoff_index}.")
    
    # results dictionary
    results_dictonary = {}
    
    # dictonary to store ground truths, predictions, and probablites from the test set
    outcome_dic = {}
    
    # dictonary for feature importance
    feature_importance_dic = {}

    os.makedirs(algorithm_folder, exist_ok=True)  # Create folder for logfile

    # create a log file for the outcome testing
    log_filename = os.path.join(algorithm_folder, "logfile.txt")
    f = open(log_filename, "w", encoding="utf-8")
    
    for outcome in outcomes:

        f.write("_____________________________________________________________________________________________________")
        f.write("\nOutcome: %s"% outcome)

        # create the train/test set for the specifc input variables and specific outcome.
        if train_data_raw is not None:
            input_train = train_data_raw[input_columns]
            train_data = pd.concat([input_train, train_data_raw[[outcome]]], axis=1)
        
        input_test = test_data_raw[input_columns]
        test_data = pd.concat([input_test, test_data_raw[[outcome]]], axis=1)
        
        if train_data_raw is not None:
            y_train = train_data[outcome]
            print(f'Training Data: {train_data}')

        y_test = test_data[outcome]
        print(f'Training Data: {test_data}')

        f.write("\nTraining Data Length: %s"% len(test_data))
        
        # Count positives and negatives
        positives = np.sum(y_test == 1)  # Count instances of 1
        negatives = np.sum(y_test == 0)  # Count instances of 0
        
        print("Postive Count (on training set):", positives)
        print("Negative Count (on training set):", negatives)

        #check_same_feature_set(train_data, test_data)
        print(test_data[outcome].value_counts())
            
        # create a states able for metric on the test set
        #res, res_array = full_roc_curve(y_test.to_numpy(), y_pred.to_numpy())
            
        try:
            # get the model
            model = models[outcome]['Model']
            print(f'Models Label: {model.label}')
            f.write("\nModel: %s"% model)
            #model = models
                
            outcome_dic[outcome] = {}
                
            #y_pred = model.predict(test_data.drop(columns=[outcome])) # get the predictions for test set
            y_proba = model.predict_proba(test_data.drop(columns=[outcome])) # Prediction Probabilities for test set
            evaluation = model.evaluate(test_data, silent=True)
                
            # save ground truths, predictions, and probabilities
            outcome_dic[outcome]['Ground Truths'] = y_test
            #outcome_dic[outcome]['Predictions'] = y_pred
            outcome_dic[outcome]['Probability Scores'] = y_proba
        
            if train_data_raw is not None:
                # create a states able for metric on the train set
                #y_pred_train = model.predict(train_data.drop(columns=[outcome]))
                y_proba_train = model.predict_proba(train_data.drop(columns=[outcome]))
                res_train, res_array_train = full_roc_curve(y_train.to_numpy(), y_proba_train[1].to_numpy(), index=cutoff_index)
                evaluation['AUROC Score (Train)'] = res_train['auc']
                evaluation['AUROC CI Low (Train)'] = res_train['auc_cilow']
                evaluation['AUROC CI High (Train)'] = res_train['auc_cihigh']
                evaluation['P (Train)'] = res_train['P'].astype(float)
                evaluation['N (Train)'] = res_train['N'].astype(float)
                evaluation['TP (Train)'] = res_train['TP']
                evaluation['FP (Train)'] = res_train['FP']
                evaluation['TN (Train)'] = res_train['TN']
                evaluation['FN (Train)'] = res_train['FN']
                
            # create a states able for metric on the test set
            res, res_array = full_roc_curve(y_test.to_numpy(), y_proba[1].to_numpy(), index=cutoff_index)
            print("Results Array: ", res)
            f.write("\nResults Array: %s"% res)
            evaluation['TPR'] = res['tpr']
            evaluation['TNR'] = res['tnr']
            evaluation['FPR'] = res['fpr']
            evaluation['FNR'] = res['fnr']
            evaluation['PPV'] = res['ppv']
            evaluation['NPV'] = res['npv']
                
            evaluation['AUROC Score'] = res['auc']
            evaluation['AUROC CI Low'] = res['auc_cilow']
            evaluation['AUROC CI High'] = res['auc_cihigh']
            evaluation['cutoff type'] = cutoff_index
            evaluation['cutoff'] = res['cutoff_mcc'] if cutoff_index=='mcc' else (res['cutoff_ji'] if cutoff_index=='ji' else (res['cutoff_f1'] if cutoff_index=='f1' else res['cutoff_youden']))
                

            y_pred = [1 if p >= evaluation['cutoff'] else 0 for p in y_proba[1]] # predict with test set
            test_acc = accuracy_score(y_test, y_pred) # test accuracy
            print("Test Accuracy:", (test_acc * 100))
            evaluation['Test Accuracy'] = test_acc
                
            # Extract TP, FP, TN, FN
            evaluation['TP'] = res['TP']
            evaluation['FP'] = res['FP']
            evaluation['TN'] = res['TN']
            evaluation['FN'] = res['FN']
                
            evaluation['P'] = res['P'].astype(float)
            evaluation['N'] = res['N'].astype(float)
                
            evaluation['precision'] = res['precision']
            evaluation['recall'] = res['recall']
            evaluation['f1 score'] = res['f1 score']
                
            evaluation['Ground Truths'] = y_test.tolist()
            evaluation['Predictions'] = y_pred
            evaluation['Probability Scores'] = y_proba.values.tolist()
                
            print(evaluation)
        except:
            print(f"Unable to make prediction for {outcome}.")
            #print(f'Its value count is {y_test.value_counts()}')
            continue

        print("Feature Importance....")
        f.write("Feature Importance....")
        original_features = model.features(feature_stage='original')

        # Get feature importance
        feature_importance_df = model.feature_importance(test_data, features=original_features, time_limit=500)
        feature_importance_df = feature_importance_df.reset_index(names='feature')
        print("Feature Importance Array: ",feature_importance_df)

        f.write("\nFeature Importance Array: %s"% feature_importance_df)

        # Sort by importance
        feature_importance_df_sorted = feature_importance_df.sort_values('importance', ascending=False)
        print(feature_importance_df_sorted)
        feature_importance_dic[outcome] = feature_importance_df_sorted

        # Get top 10 most important feature names
        top_features = feature_importance_df_sorted['feature'].head(10).tolist()
        #all_features = feature_importance_df_sorted['feature'].tolist()
        f.write("\nTop 10 important features: %s" % top_features)

        results_dictonary[outcome] = {}
        model.leaderboard(test_data)
        print(f'Best Model for {outcome} is {model.model_best}.')
        results_dictonary[outcome]['best_model'] = model.model_best
        results_dictonary[outcome]['evaluation'] = evaluation
        results_dictonary[outcome]['leaderboard'] = model.leaderboard(test_data).to_dict(orient='records')
        try:
            results_dictonary[outcome]['feature importance'] = feature_importance_df_sorted.to_dict(orient='records')
            #results_dictonary[outcome]['top features'] = all_features
        except:
            print(f"Unable to make the feature importance table for {outcome}.")

        print(f'Testing for {outcome} is complete.')
        f.write(f'Testing for {outcome} is complete.')
        print("_________________________________________________________")
    
    # Write the ending message
    #placeholder.empty()  # Clears the output
    f.write("Model Testing is Complete!")
    print("Model Testing is Complete!")
    f.close()
    
    return results_dictonary, outcome_dic, feature_importance_dic


def auto_gulon_worker(job_id, client_name, exp_name, models_dic, model_type, test_set, data_name_test, uploaded_test_set, test_set_name, input_variables, selected_outcomes, train_set, threshold_type):
    # connect to database
    #client = MongoClient('10.14.1.12', 27017)
    client = MongoClient(client_name, 27017)

    # create the database if it does not already exists
    db = client.machine_learning_database
    # create tables for models in the databse
    models = db.models
    # create the results if it does not already exists
    results = db.results
    # create the results if it does not already exists
    datasets = db.datasets
    # create the jobs if it does not already exists
    jobs = db.jobs
    # get all unique exp. names from results collection
    exp_names = db.models.distinct("exp_name", {"type": "AutoGulon"})
    # get all training data names from database
    data_names_train = db.datasets.distinct("data_name", {"type": "Train"})
    # get all testing data names from database
    data_names_list_test = db.datasets.distinct("data_name", {"type": "Test"})

    # create a jobs dictonary to store in the MongoDB database
    job_doc = {
        "job_id": job_id,
        "job_type": "Testing (AutoGulon)",
        "exp_name": exp_name,
        "status": "queued",
        "message": "Waiting to start...",
        "created_at": datetime.now(),
        "last_updated": datetime.now()
    }
    db.jobs.insert_one(job_doc)

    def update_status(state, message): # inner function to update the status of the current job
        db.jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": state,
                "message": message,
                "last_updated": datetime.now()
            }})
        
    def update_status_when_done(state, message): # inner function to update the status when it is finsihed
        completed_time = datetime.now()
        expires_at = completed_time + timedelta(days=1)
        db.jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": state,
                "message": message,
                "last_updated": completed_time,
                "completed_at": completed_time,
                "expires_at": expires_at
            }})
        
    def heartbeat():
        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "heartbeat": datetime.now(),
                "last_updated": datetime.now()
            }}
        )

    stop_heartbeat = False
    def heartbeat_loop():
        while not stop_heartbeat:
            heartbeat()
            time.sleep(30)

    heartbeat_thread = threading.Thread(
        target=heartbeat_loop,
        daemon=True
    )
    heartbeat_thread.start()
        
    
    def handle_termination(signum, frame):
        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "Stopped",
                "message": "Job was terminated unexpectedly.",
                "last_updated": datetime.now(),
                "expires_at": datetime.now() + timedelta(days=1)
            }}
        )
        sys.exit(1)

    signal.signal(signal.SIGTERM, handle_termination)
    signal.signal(signal.SIGINT, handle_termination)

    try:
        print("Starting Experiment...")

        update_status("running", "Starting saving test data...")
        #save the test set if it hasn't already
        if data_name_test == None and uploaded_test_set.name not in data_names_list_test:
            print(f"Testing Dataset is saving in the database")
            # create a list of ML exp.'s that the dataset was used on
            exp_list = [exp_name]
            #get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
            # save test set in data folder and database
            os.makedirs("Data Sets", exist_ok=True)
            save_data(uploaded_test_set.name, test_set, os.path.join("Data Sets", uploaded_test_set.name))
            dataset_test = {
                    "data_name": uploaded_test_set.name,
                    "type": "Test",
                    "time_saved": current_time,
                    "data_path": os.path.join("Data Sets", uploaded_test_set.name),
                    "exps used": exp_list
            }
            datasets.insert_one(dataset_test)
        else:
            print(f"Testing Dataset of the same name is already in the database. Will be overwritten in the database")
            test_name = uploaded_test_set.name if data_name_test == None else data_name_test
            save_data(test_name, test_set, os.path.join("Data Sets", test_name))
            # update the dataset in datbase to trackdown the list of ML exps the set was used on
            dataset = datasets.find_one({"data_name": test_name, "type": "Test"})
            # Get the current list of experiments or initialize it if not present
            exp_list = dataset.get("exps used", [])
            # Add the current project name if it's not already in the list
            if exp_name not in exp_list:
                exp_list.append(exp_name)

            datasets.update_one(
                {"data_name": test_name, "type": "Test"}, # Filter condition
                {"$set": { "exps used": exp_list }} # Update operation
            )

        algorithm_folder = os.path.join("Results", exp_name, test_set_name)
        os.makedirs(algorithm_folder, exist_ok=True)  # Create folder for algorithm results

        update_status("running", "Starting testing...")
        # call the function to test models
        results_dictonary, outcome_dic, feature_importance_dic = test_model(models_dic, test_set, input_variables, selected_outcomes, train_data_raw=train_set, cutoff_index=threshold_type, algorithm_folder=algorithm_folder)
        
        update_status("running", "Testing Done, Now saving results...")

        path_name = f"{algorithm_folder}/{exp_name}_results.joblib"
        joblib.dump(results_dictonary, path_name)

        # generate the feature importance charts
        chart_name = f"{algorithm_folder}/Feature Importance"
        plot_feature_importance(exp_name, feature_importance_dic, chart_name)

        # generate the results table
        results_df = generate_results_table(results_dictonary, selected_outcomes)
        table_name = f"{algorithm_folder}/{exp_name}_results.xlsx"

        results_df.to_excel(table_name, index=False, engine='openpyxl')

        results_df.to_excel(table_name, index=False)
        results_df.to_excel(table_name, index=False)
        expand_cell_excel(table_name)
        wrap_text_excel(table_name)
        grid_excel(table_name)

        # Convert results_df to list of dictionaries
        results_dic = results_df.to_dict(orient='records')
        # get the current time
        current_datetime = datetime.now()
        current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
        # get test dataset name
        test_name = uploaded_test_set.name if data_name_test == None else data_name_test
        #st.write(results_dictonary)
        try:
            result = {
                "exp_name": exp_name,
                "type": model_type,
                "test set": test_set_name,
                "threshold used": threshold_type,
                "results_dic": results_dictonary,
                "results_table": results_dic,
                'dataset used': test_name,
                "time_created": current_time
            }

            results.insert_one(result) # insert one dictonary
        except:
            print("Results size is too large. Will save filepaths instead.")

            result = {
                "exp_name": exp_name,
                "type": model_type,
                "test set": test_set_name,
                "threshold used": threshold_type,
                "results_dic": path_name,
                "results_table": results_dic,
                'dataset used': test_name,
                "time_created": current_time
            }
            results.insert_one(result) # insert one dictonary

        update_status_when_done("Completed", "Training complete")

    except Exception as e:
        error_message = str(e)
        stack_trace = traceback.format_exc()

        print(f"Error Occured. Experiment '{exp_name}' failed!")
        traceback.print_exc()

        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "Failed",
                "message": error_message,
                "error_trace": stack_trace,
                "last_updated": datetime.now(),
                "expires_at": datetime.now() + timedelta(days=1)
            }}
        )

    finally:
        stop_heartbeat = True
        heartbeat_thread.join(timeout=1)



