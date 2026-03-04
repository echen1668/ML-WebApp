from pathlib import Path
import time
import io
import pandas as pd
import numpy as np
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn import metrics
import time
import psutil
import autogluon
import pandas as pd
import numpy as np
import json
import sklearn as scikit_learn
import seaborn as sns
import matplotlib.pyplot as plt
#from scipy import interp
from scipy.stats import norm
from scipy.special import ndtri
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import magic 
import pickle
import random 
from random import randint
from random import uniform
from scipy import stats
import ast
from pathlib import Path
import joblib as joblib
from joblib import dump, load
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from sklearn.model_selection import cross_validate
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import RocCurveDisplay
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import PrecisionRecallDisplay
from sklearn.metrics import auc
from sklearn.utils import resample
from sklearn.model_selection import KFold
from sklearn.model_selection import StratifiedKFold, RepeatedStratifiedKFold
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import VarianceThreshold
from sklearn.model_selection import RandomizedSearchCV, GridSearchCV
from sklearn.calibration import CalibratedClassifierCV
from sklearn.calibration import calibration_curve, CalibrationDisplay
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score
from sklearn.metrics import classification_report
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import label_binarize
from sklearn.metrics import confusion_matrix
#from scipy import interp
from scipy.stats import norm
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
import random 
from random import randint
from random import uniform
from scipy import stats
import mrmr
from mrmr import mrmr_classif
import xgboost
import catboost
import shap
from scipy import stats
import os
import signal
import sys
import joblib as joblib
from joblib import dump, load
import json
import threading
import traceback
import tkinter as tk
from tkinter import *
from autogluon.tabular import TabularDataset, TabularPredictor
import mimetypes
from difflib import SequenceMatcher
#np.random.seed(1000)
rstate = 12
from datetime import datetime, timedelta
# import module
import pprint
import pymongo
from pymongo import MongoClient
from streamlit_cookies_manager import EncryptedCookieManager
from Multi_Outcome_Classification_tools import multi_outcome_hyperparameter_binary, multi_outcome_hyperparameter_binary_train_and_test, multi_outcome_cv
from Common_Tools import generate_configuration_file, generate_shap_table, generate_configuration_template, generate_results_table, generate_congfig_file, get_avg_results_dic, wrap_text_excel, expand_cell_excel, grid_excel, generate_all_idx_files, upload_data, load_data, save_data, data_prep, data_prep_train_set, parse_exp_multi_outcomes, setup_multioutcome_binary, refine_binary_outcomes, generate_joblib_model
from roctools import full_roc_curve, plot_roc_curve




# run training
def project(job_id, client_name, configuration_dic, data_sets, unique_value_threshold, training_method, project_name):
    print("Currently training the models...")
    # connect to database
    client = MongoClient(client_name, 27017)
    # create the database if it does not already exists
    db = client.machine_learning_database
    # create tables for models in the databse
    models = db.models
    # create the results if it does not already exists
    jobs = db.jobs
    db.jobs.create_index("expires_at", expireAfterSeconds=0)
    # create the results if it does not already exists
    results = db.results
    # create the results if it does not already exists
    datasets = db.datasets
    # get all unique exp. names from results collection
    exp_names_n = db.models.distinct("exp_name")
    exp_names_cv = db.results.distinct("exp_name", {"type": "Native-CV"})
    exp_names = exp_names_n + exp_names_cv
    # get all training data names from database
    data_names_train = db.datasets.distinct("data_name", {"type": "Train"})
    # get all testing data names from database
    data_names_list_test = db.datasets.distinct("data_name", {"type": "Test"})

    # create a jobs dictonary to store in the MongoDB database
    job_doc = {
        "job_id": job_id,
        "job_type": f"Training (Sklearn - {training_method})",
        "exp_name": project_name,
        "status": "queued",
        "message": "Waiting to start...",
        "created_at": datetime.now(),
        "last_updated": datetime.now()
    }
    db.jobs.insert_one(job_doc)
        
    def update_status(state, message): # inner function to update the status of the current job
        db.jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": state,
                "message": message,
                "last_updated": datetime.now()
            }})
        
    def update_status_when_done(state, message): # inner function to update the status when it is finsihed
        completed_time = datetime.now()
        expires_at = completed_time + timedelta(days=1)
        db.jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": state,
                "message": message,
                "last_updated": completed_time,
                "completed_at": completed_time,
                "expires_at": expires_at
            }})
        
        
    def heartbeat():
        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "heartbeat": datetime.now(),
                "last_updated": datetime.now()
            }}
        )

    stop_heartbeat = False
    def heartbeat_loop():
        while not stop_heartbeat:
            heartbeat()
            time.sleep(30)

    heartbeat_thread = threading.Thread(
        target=heartbeat_loop,
        daemon=True
    )
    heartbeat_thread.start()
        
    
    def handle_termination(signum, frame):
        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "Stopped",
                "message": "Job was terminated unexpectedly.",
                "last_updated": datetime.now(),
                "expires_at": datetime.now() + timedelta(days=1)
            }}
        )
        sys.exit(1)

    signal.signal(signal.SIGTERM, handle_termination)
    signal.signal(signal.SIGINT, handle_termination)


    try:

        print("Starting Experiment...")

        update_status("running", "Saving the training data...")

        # --- 1. Setup Folders ---
        # get name of the project as part of file name
        project_name = list(configuration_dic.keys())[0]
        project_folder = os.path.join("Models", project_name)
        os.makedirs(project_folder, exist_ok=True)  # Create folder for algorithm results

        train_set = data_sets["Training Set"]
        index_set = data_sets["Index Set"]
        #test_sets = data_sets["Testing Set"]

        # save the data sets
        #save_data(train_set['Name'], train_set['Data'], os.path.join(project_folder, train_set['Name']))
        save_data(index_set['Name'], index_set['Data'], os.path.join(project_folder, index_set['Name']))

        if train_set['Name'] not in data_names_train:
            print(f"Training Dataset {train_set['Name']} is saving in the database")
            # create a list of ML exp.'s that the dataset was used on
            exp_list = [project_name]
            # get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
            # save train set into data folder and database
            os.makedirs("Data Sets", exist_ok=True)
            save_data(train_set['Name'], train_set['Data'], os.path.join("Data Sets", train_set['Name']))
            dataset_train = {
                    "data_name": train_set['Name'],
                    "type": "Train",
                    "time_saved": current_time,
                    "data_path": os.path.join("Data Sets", train_set['Name']),
                    "exps used": exp_list
            }
            datasets.insert_one(dataset_train)
        else:
            save_data(train_set['Name'], train_set['Data'], os.path.join("Data Sets", train_set['Name']))
            # update the dataset in datbase to trackdown the list of ML exps the set was used on
            dataset = datasets.find_one({"data_name": train_set['Name'], "type": "Train"})
            # Get the current list of experiments or initialize it if not present
            exp_list = dataset.get("exps used", [])
            # Add the current project name if it's not already in the list
            if project_name not in exp_list:
                exp_list.append(project_name)

            datasets.update_one(
                {"data_name": train_set['Name'], "type": "Train"}, # Filter condition
                {"$set": { "exps used": exp_list }} # Update operation
            )

        # get proper data name
        if train_set['Name'].endswith('.xlsx'):
            data_name = train_set["Name"][:-len('.xlsx')]
        elif train_set["Name"].endswith('.csv'):
            data_name = train_set["Name"][:-len('.csv')]
        else:
            data_name = train_set["Name"]

        # get input and ouput columns
        input_cols, label_cols, categorical_cols, numeric_cols = parse_exp_multi_outcomes(train_set["Data"], index_set["Data"], unique_value_threshold=unique_value_threshold)
        input_cols_og = input_cols.copy() # keep the origional input list

        threshold_type = configuration_dic[project_name]['threshold_type']
        if threshold_type not in ['youden', 'mcc', 'ji', 'f1']: # if  threshold_type not a valid value, the program will automaticlly default to youden
            print("Threshold Type value is not valid, will default to youden.")
            threshold_type = 'youden'

        df_train = refine_binary_outcomes(train_set["Data"], label_cols)

        algorithms = [] # list of algorithims
        results_dictonary = {}
        for experiment in configuration_dic[project_name]:
            experiment_name = experiment
            if experiment_name in ["train_set", "test_sets", "threshold_type", "exp_type"]:
                continue

            # --- 2. Prep the data ---
            print(f"Experiment Name: {experiment_name} had started training...")
            
            print(f"Preping the data for {experiment_name}...")
            set_up = configuration_dic[project_name][experiment_name]
            options, algorithm, param_vals = setup_multioutcome_binary(set_up, experiment_name, project_folder)

            algorithms.append(algorithm)
            results_dictonary[algorithm] = {}
            update_status("running", f"Training models for {experiment_name} for {algorithm}...")

            # refine the training set before model training
            df_train_refined, input_cols, label_cols, encoder, encoded_cols, qt = data_prep(df_train, input_cols, label_cols, numeric_cols, categorical_cols, options)
            # save any preprocessing steps
            algorithm_folder = os.path.join(project_folder, experiment_name)
            os.makedirs(algorithm_folder, exist_ok=True)  # Create folder for algorithm preprocessing steps

            if encoder != None:
                encoder_name = os.path.join(algorithm_folder, algorithm + "_encoder.joblib")
                joblib.dump(encoder, encoder_name)
                encoded_cols_name = os.path.join(algorithm_folder, algorithm + "_encoded_cols.joblib")
                joblib.dump(encoded_cols, encoded_cols_name)

            if qt != None:
                qt_name = os.path.join(algorithm_folder, algorithm + "_qt.joblib")
                joblib.dump(qt, qt_name)

            # --- 3. Loop Through Algorithms and Execute Workflows ---
            print(f"Training has offically started for {experiment_name}! ... This may take a while.")

            if training_method == "Train Whole Set":
                multi_outcome_hyperparameter_binary(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, options, algorithm, param_vals, experiment_name, project_folder)
                
            elif training_method == "Train/Test Split":
                algo_dictonary = multi_outcome_hyperparameter_binary_train_and_test(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, threshold_type, options, algorithm, param_vals, experiment_name, project_folder, project_name, data_name)
                results_dictonary[algorithm] = algo_dictonary
            
            elif training_method == "Cross-Validation":
                algo_dictonary = multi_outcome_cv(df_train_refined, input_cols, label_cols, numeric_cols, categorical_cols, threshold_type, options, algorithm, param_vals, experiment_name, project_folder, project_name)
                results_dictonary[algorithm] = algo_dictonary

            print(f"Successfully completed training for **{experiment_name}**.")
        
        
        print(f"**Successfully completed training for all experiments!**")
        
        # --- 4. Finalize Experiment ---
        print("Finalizing experiment: saving metadata and final reports...")
        update_status("running", "Training Done. Saving the models...")

        if training_method in ["Train/Test Split", "Train Whole Set"]:
            models_dic, _ = generate_joblib_model(project_folder)
            model_absolute_path = os.path.join(project_folder)

            pathway_name = os.path.join(model_absolute_path, project_name + "_models.joblib")
            joblib.dump(models_dic, pathway_name)

            # get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
            
            model = {
                "exp_name": project_name,
                "type": "Native",
                "model_path": pathway_name,
                "algorithms": algorithms,
                "input variables": input_cols,
                "input variables (original)": input_cols_og,
                'outcomes': label_cols,
                "configuration": configuration_dic,
                "train_data": train_set["Name"],
                "time_created": current_time
            }

            models.insert_one(model) # insert one dictonary 
        
        if training_method in ["Train/Test Split", "Cross-Validation"]:

            # save the results into database
            results_folder = os.path.join("Results", project_name, data_name) if training_method=="Train/Test Split" else os.path.join("Results", project_name)
            os.makedirs(results_folder, exist_ok=True)  # Create folder for algorithm results if it doesn't exists yet
            filename = os.path.join(results_folder, "metadata.txt")
            f = open(filename, "w", encoding="utf-8")
            f.write("\nExp Name: %s"% project_name)
            f.write("\nInput Columns: %s"% input_cols_og)
            f.write("\nOutput Columns: %s"% label_cols)
            f.write("\nAlgorithms: %s"% algorithms)
            f.write("\nConfiguration: %s"% configuration_dic)
            f.close()
            

            training_type = "Native" if training_method=="Train/Test Split" else "Native-CV"

            final_results_dic = get_avg_results_dic(results_dictonary, override=True) if training_method=="Cross-Validation" else results_dictonary

            path_name = os.path.join(results_folder, f"{project_name}_results.joblib")
            joblib.dump(final_results_dic, path_name)

            # generate the results table
            results_df = generate_results_table(final_results_dic)
            table_name = os.path.join(results_folder, f"{project_name}_results.xlsx")

            results_df.to_excel(table_name, index=False, engine='openpyxl')

            results_df.to_excel(table_name, index=False)
            results_df.to_excel(table_name, index=False)
            expand_cell_excel(table_name)
            wrap_text_excel(table_name)
            grid_excel(table_name)

            # Convert results_df to list of dictionaries
            results_dic = results_df.to_dict(orient='records')

            # get the current time
            current_datetime = datetime.now()
            current_time = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

            try:
                result = {
                    "exp_name": project_name,
                    "type": training_type,
                    "test set": data_name,
                    "threshold used": threshold_type,
                    "results_dic": final_results_dic,
                    "results_table": results_dic,
                    'dataset used': train_set['Name'],
                    "algorithms": algorithms,
                    "input variables": input_cols,
                    "input variables (original)": input_cols_og,
                    'outcomes': label_cols,
                    "time_created": current_time
                }

                if training_method == "Cross-Validation": # add configuration if the experiment is cross validation
                    result['configuration'] = configuration_dic
                    # create SHAP Values table
                    _, shap_table_path = generate_shap_table(results_dictonary, project_name) # get the path for the SHAP table file
                    result['SHAP Table'] = shap_table_path

                results.insert_one(result) # insert one dictonary
            except:
                print("Results size is too large. Will save filepaths instead.")

                result = {
                    "exp_name": project_name,
                    "type": training_type,
                    "test set": data_name,
                    "threshold used": threshold_type,
                    "results_dic": path_name,
                    "results_table": results_dic,
                    'dataset used': train_set['Name'],
                    "algorithms": algorithms,
                    "input variables": input_cols,
                    "input variables (original)": input_cols_og,
                    'outcomes': label_cols,
                    "time_created": current_time
                }

                if training_method == "Cross-Validation": # add configuration if the experiment is cross validation
                    result['configuration'] = configuration_dic
                    # create SHAP Values table
                    _, shap_table_path = generate_shap_table(results_dictonary, project_name) # get the path for the SHAP table file
                    result['SHAP Table'] = shap_table_path

                results.insert_one(result) # insert one dictonary
        
        print(f"Experiment '{project_name}' completed successfully!")
        update_status_when_done("Completed", "Training complete")
        
    except Exception as e:
        error_message = str(e)
        stack_trace = traceback.format_exc()

        print(f"Error Occured. Experiment '{project_name}' failed!")
        traceback.print_exc()

        jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "Failed",
                "message": error_message,
                "error_trace": stack_trace,
                "last_updated": datetime.now(),
                "expires_at": datetime.now() + timedelta(days=1)
            }}
        )

    finally:
        stop_heartbeat = True
        heartbeat_thread.join(timeout=1)

